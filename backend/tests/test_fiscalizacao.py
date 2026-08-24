"""Módulo Fiscalização — turnos, partidas, eventos e fechamento (Blocos A/B/C).

Mesmo padrão de test_portaria.py: SQLite em memória com ATTACH DATABASE
para os schemas `fiscalizacao` e `portaria` (isolados do `main`, igual à
separação real do Postgres), sem conftest.py. O gate RBAC (`exige()`) usa
`vw_acesso_efetivo`, view Postgres que não existe em SQLite — por isso as
dependências de permissão dos routers são sobrescritas diretamente.

Os testes que provam o comportamento ERRADO são os que valem (§8 do prompt):
um teste que só confirma o caminho feliz não prova nada.

⛔ Nenhum dado pessoal real — RE, prefixo e nome fictícios.
"""
import sqlite3
import typing
import uuid as _uuid_mod
from datetime import date, datetime, time, timedelta, timezone
from uuid import UUID, uuid4

import os
from pathlib import Path

import pytest
from fastapi import HTTPException, status
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event, select
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.core.config import FUSO_OPERACAO
from app.core.database import Base, get_db
from app.main import app
from app.models.cadastro import Funcao, Funcionario, FuncionarioFuncao
from app.models.catalogos import Linha
from app.models.fiscalizacao import (
    Baita, EventoTurno, LinhaCoordenador, ObservacaoTurno, PartidaProgramada, Ponto, PontoLinha,
    RegistroPartida, Turno, TurnoLinha,
)
from app.models.enums import SetorEnum
from app.models.portaria import RecolhidaAnormal
from app.routers import fiscalizacao as fiscalizacao_router_mod
from app.services.fechamento_fiscal import calcular_fechamento_linha, montar_fechamento
from app.services.importacao_escala_fiscal import _classificar_periodo, _parse_aba, _parse_limites_periodo

sqlite3.register_adapter(_uuid_mod.UUID, lambda u: u.hex)


def _dependency_de(annotated_type):
    for arg in typing.get_args(annotated_type)[1:]:
        dependency = getattr(arg, "dependency", None)
        if dependency is not None:
            return dependency
    raise RuntimeError("Depends não encontrado no tipo anotado")


_FISCAL_A = Funcionario(id=uuid4(), re="70001", nome="Fiscal A")
_FISCAL_B = Funcionario(id=uuid4(), re="70002", nome="Fiscal B")
_COORDENADOR = Funcionario(id=uuid4(), re="70003", nome="Coordenador Teste")
_ADMIN = Funcionario(id=uuid4(), re="70004", nome="Admin Teste")
_COORDENADOR_B = Funcionario(id=uuid4(), re="70005", nome="Coordenador Teste B")

_TABELAS = [
    Funcionario.__table__, Funcao.__table__, FuncionarioFuncao.__table__,
    Ponto.__table__, PontoLinha.__table__, Turno.__table__, TurnoLinha.__table__,
    PartidaProgramada.__table__, RegistroPartida.__table__, EventoTurno.__table__, ObservacaoTurno.__table__,
    Baita.__table__, RecolhidaAnormal.__table__, LinhaCoordenador.__table__, Linha.__table__,
]

_PERMISSOES = {
    "FISCAL": {
        "leitura": True, "escrita": True, "leitura_painel": False, "escrita_painel": False,
        "escrita_escala": False,
    },
    "COORDENADOR": {
        "leitura": True, "escrita": False, "leitura_painel": True, "escrita_painel": True,
        "escrita_escala": True,
    },
    "ADMIN": {
        "leitura": True, "escrita": True, "leitura_painel": True, "escrita_painel": True,
        "escrita_escala": True,
    },
}
_USUARIOS_PADRAO = {"FISCAL": _FISCAL_A, "COORDENADOR": _COORDENADOR, "ADMIN": _ADMIN}


@pytest.fixture
def ambiente():
    engine = create_engine(
        "sqlite:///:memory:",
        poolclass=StaticPool,
        connect_args={"check_same_thread": False},
    )

    @event.listens_for(engine, "connect")
    def _attach(dbapi_conn, _):
        dbapi_conn.execute("ATTACH DATABASE ':memory:' AS fiscalizacao")
        dbapi_conn.execute("ATTACH DATABASE ':memory:' AS portaria")
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine, tables=_TABELAS)

    with Session(engine) as setup:
        for f in (_FISCAL_A, _FISCAL_B, _COORDENADOR, _ADMIN, _COORDENADOR_B):
            setup.add(Funcionario(id=f.id, re=f.re, nome=f.nome, status="ATIVO"))
        setup.add(Ponto(codigo="PQ_TESTE", nome="Ponto Teste", terminal="TP", ativo=True))
        setup.add(PontoLinha(ponto_codigo="PQ_TESTE", linha_codigo="1726-10", ativo=True))
        setup.add(Linha(id=uuid4(), codigo="1726-10", nome="Linha 1726-10", setor=SetorEnum.E2, ativa=True))
        setup.add(Linha(id=uuid4(), codigo="2032-10", nome="Linha 2032-10", setor=SetorEnum.AR2, ativa=True))
        setup.add(Linha(id=uuid4(), codigo="9999-10", nome="Linha 9999-10 (inativa)", setor=SetorEnum.E2, ativa=False))
        setup.commit()

    def _get_db_teste():
        db = Session(engine)
        try:
            yield db
        finally:
            db.close()

    deps = {
        "leitura": _dependency_de(fiscalizacao_router_mod.LeituraFiscalizacao),
        "escrita": _dependency_de(fiscalizacao_router_mod.EscritaFiscalizacao),
        "leitura_painel": _dependency_de(fiscalizacao_router_mod.LeituraPainel),
        "escrita_painel": _dependency_de(fiscalizacao_router_mod.EscritaPainel),
        "escrita_escala": _dependency_de(fiscalizacao_router_mod.EscritaEscala),
    }

    app.dependency_overrides[get_db] = _get_db_teste

    yield {"engine": engine, "http": TestClient(app), **deps}

    for dep in deps.values():
        app.dependency_overrides.pop(dep, None)
    app.dependency_overrides.pop(get_db, None)


def _negar():
    def _dep():
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="Sem permissão (fake, teste)")
    return _dep


def _permitir(usuario):
    def _dep():
        return usuario
    return _dep


def _como(ambiente, papel, usuario=None):
    usuario = usuario or _USUARIOS_PADRAO[papel]
    for chave, permitido in _PERMISSOES[papel].items():
        dep = ambiente[chave]
        ambiente["http"].app.dependency_overrides[dep] = _permitir(usuario) if permitido else _negar()
    return usuario


def _criar_turno(ambiente, *, funcionario=_FISCAL_A, ponto_codigo="PQ_TESTE", terminal="TP",
                  periodo="1", data_referencia=None, tipo_dia="UTIL", status_="ABERTO",
                  linhas=("1726-10",), **campos) -> Turno:
    data_referencia = data_referencia or datetime.now(FUSO_OPERACAO).date()
    turno = Turno(
        id=uuid4(), funcionario_id=funcionario.id, fiscal_re=funcionario.re,
        ponto_codigo=ponto_codigo, terminal=terminal, periodo=periodo,
        data_referencia=data_referencia, tipo_dia=tipo_dia, status=status_,
        **campos,
    )
    with Session(ambiente["engine"]) as db:
        db.add(turno)
        db.flush()
        for linha_codigo in linhas:
            db.add(TurnoLinha(turno_id=turno.id, linha_codigo=linha_codigo))
        db.commit()
        db.refresh(turno)
        db.expunge(turno)
    return turno


def _criar_partida_programada(ambiente, **campos) -> None:
    with Session(ambiente["engine"]) as db:
        db.add(PartidaProgramada(id=uuid4(), **campos))
        db.commit()


def _criar_registro(ambiente, **campos) -> UUID:
    registro_id = uuid4()
    with Session(ambiente["engine"]) as db:
        db.add(RegistroPartida(id=registro_id, **campos))
        db.commit()
    return registro_id


def _criar_evento(ambiente, **campos) -> None:
    with Session(ambiente["engine"]) as db:
        db.add(EventoTurno(id=uuid4(), **campos))
        db.commit()


class _DatetimeFixo(datetime):
    """Mesmo padrão de test_portaria.py — congela `datetime.now()` do
    módulo do router para os testes de estado derivado (D7) não ficarem
    reféns do relógio da máquina."""
    _fixo_utc: datetime = None

    @classmethod
    def now(cls, tz=None):
        if tz is None:
            return cls._fixo_utc.replace(tzinfo=None)
        return cls._fixo_utc.astimezone(tz)


def _congelar_relogio(monkeypatch, momento_sp: datetime) -> None:
    fixo = type("_DatetimeFixoInstancia", (_DatetimeFixo,), {"_fixo_utc": momento_sp.astimezone(timezone.utc)})
    monkeypatch.setattr(fiscalizacao_router_mod, "datetime", fixo)


# ============================================================================
# CATÁLOGO DE LINHAS — GET /fiscalizacao/catalogo/linhas
# ============================================================================

def test_catalogo_linhas_devolve_so_ativas_ordenadas(ambiente):
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].get("/fiscalizacao/catalogo/linhas")
    assert resp.status_code == 200, resp.text
    codigos = [item["codigo"] for item in resp.json()]
    assert codigos == ["1726-10", "2032-10"]
    assert "9999-10" not in codigos


def test_catalogo_linhas_incluir_inativas_devolve_tambem_a_inativa(ambiente):
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].get("/fiscalizacao/catalogo/linhas?incluir_inativas=true")
    assert resp.status_code == 200, resp.text
    codigos = [item["codigo"] for item in resp.json()]
    assert "9999-10" in codigos


def test_fiscal_sem_painel_le_catalogo_de_linhas(ambiente):
    # O endpoint existe justamente pra isto: FISCAL não tem
    # fiscalizacao_painel, mas tem leitura em fiscalizacao — o catálogo
    # precisa estar acessível pra ele mesmo assim (D37/D38 dependem disto).
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].get("/fiscalizacao/catalogo/linhas")
    assert resp.status_code == 200, resp.text


# ============================================================================
# 1 — PERDIDA sem motivo → recusado
# ============================================================================

def test_perdida_sem_motivo_recusado(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        "linha_codigo": "1726-10", "numero_tabela": 1, "terminal": "TP",
        "horario_programado": "08:00:00", "resultado": "PERDIDA",
    })
    assert resp.status_code == 422, resp.text


# ============================================================================
# 2 — motivo=OUTRO sem motivo_outro → recusado
# ============================================================================

def test_motivo_outro_sem_texto_recusado(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        "linha_codigo": "1726-10", "numero_tabela": 1, "terminal": "TP",
        "horario_programado": "08:00:00", "resultado": "PERDIDA", "motivo": "OUTRO",
    })
    assert resp.status_code == 422, resp.text


# ============================================================================
# 3 — motivo=RA sem prefixo → recusado (D18)
# ============================================================================

def test_motivo_ra_sem_prefixo_recusado(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        "linha_codigo": "1726-10", "numero_tabela": 1, "terminal": "TP",
        "horario_programado": "08:00:00", "resultado": "PERDIDA", "motivo": "RA",
    })
    assert resp.status_code == 422, resp.text


# ============================================================================
# 4 — 🔴 Fiscal A registrando partida no turno do fiscal B → 403
# ============================================================================

def test_fiscal_a_no_turno_do_fiscal_b_nega_403(ambiente):
    turno_b = _criar_turno(ambiente, funcionario=_FISCAL_B)
    _como(ambiente, "FISCAL", usuario=_FISCAL_A)
    resp = ambiente["http"].put(f"/fiscalizacao/turnos/{turno_b.id}/partidas", json={
        "linha_codigo": "1726-10", "numero_tabela": 1, "terminal": "TP",
        "horario_programado": "08:00:00", "resultado": "REALIZADA",
    })
    assert resp.status_code == 403, resp.text

    detalhe = ambiente["http"].get(f"/fiscalizacao/turnos/{turno_b.id}")
    assert detalhe.status_code == 403, detalhe.text


# ============================================================================
# 5 — 🔴 FISCAL chamando /fiscalizacao/painel/... → 403 (sem fiscalizacao_painel)
# ============================================================================

def test_fiscal_no_painel_nega_403(ambiente):
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].get("/fiscalizacao/painel/1726-10")
    assert resp.status_code == 403, resp.text


def test_coordenador_acessa_painel(ambiente):
    _como(ambiente, "COORDENADOR")
    resp = ambiente["http"].get("/fiscalizacao/painel/1726-10")
    assert resp.status_code == 200, resp.text
    assert resp.json()["linha_codigo"] == "1726-10"


# ============================================================================
# 6 — segundo turno ABERTO igual (funcionário/ponto/período/data) → 409
# ============================================================================

def test_segundo_turno_aberto_igual_nega_409(ambiente):
    _como(ambiente, "FISCAL")
    primeiro = ambiente["http"].post("/fiscalizacao/turnos", json={
        "ponto_codigo": "PQ_TESTE", "periodo": "1", "linhas": ["1726-10"],
    })
    assert primeiro.status_code == 201, primeiro.text

    segundo = ambiente["http"].post("/fiscalizacao/turnos", json={
        "ponto_codigo": "PQ_TESTE", "periodo": "1", "linhas": ["1726-10"],
    })
    assert segundo.status_code == 409, segundo.text


# ============================================================================
# 7 — PUT /partidas duas vezes no mesmo horário → atualiza, não duplica
# (e o evento_turno vinculado acompanha)
# ============================================================================

def test_put_partida_duas_vezes_atualiza_nao_duplica(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")

    payload_base = {
        "linha_codigo": "1726-10", "numero_tabela": 1, "terminal": "TP", "horario_programado": "08:00:00",
    }

    primeira = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        **payload_base, "resultado": "PERDIDA", "motivo": "RA", "prefixo": "1721",
    })
    assert primeira.status_code == 200, primeira.text
    registro_id = primeira.json()["id"]

    with Session(ambiente["engine"]) as db:
        registros = db.execute(select(RegistroPartida).where(RegistroPartida.turno_id == turno.id)).scalars().all()
        eventos = db.execute(select(EventoTurno).where(EventoTurno.turno_id == turno.id)).scalars().all()
    assert len(registros) == 1
    assert len(eventos) == 1
    assert eventos[0].tipo == "RA"
    assert eventos[0].registro_partida_id == UUID(registro_id)

    # Muda a resposta: agora saiu. O evento RA vinculado tem que sumir —
    # D4: resultado REALIZADA não tem contador nenhum vinculado.
    segunda = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        **payload_base, "resultado": "REALIZADA",
    })
    assert segunda.status_code == 200, segunda.text
    assert segunda.json()["id"] == registro_id  # mesmo registro, não um novo

    with Session(ambiente["engine"]) as db:
        registros = db.execute(select(RegistroPartida).where(RegistroPartida.turno_id == turno.id)).scalars().all()
        eventos = db.execute(select(EventoTurno).where(EventoTurno.turno_id == turno.id)).scalars().all()
    assert len(registros) == 1  # não duplicou
    assert len(eventos) == 0  # evento vinculado foi removido junto


# ============================================================================
# 8 — PUT /baita duas vezes → atualiza, não duplica
# ============================================================================

def test_put_baita_duas_vezes_atualiza_nao_duplica(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")

    primeira = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/baita", json={
        "linha_codigo": "1726-10", "tipo": "BAITA", "prefixo": "2704",
    })
    assert primeira.status_code == 200, primeira.text

    segunda = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/baita", json={
        "linha_codigo": "1726-10", "tipo": "BAITA", "prefixo": "2705", "motorista_re": "8201",
    })
    assert segunda.status_code == 200, segunda.text
    assert segunda.json()["prefixo"] == "2705"
    assert segunda.json()["id"] == primeira.json()["id"]

    with Session(ambiente["engine"]) as db:
        registros = db.execute(select(Baita).where(Baita.turno_id == turno.id)).scalars().all()
    assert len(registros) == 1


# ============================================================================
# 9 — ts_circular=true com saida_ts preenchido → recusado
# ============================================================================

def test_ts_circular_com_saida_ts_recusado(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/baita", json={
        "linha_codigo": "1726-10", "tipo": "BAITA", "prefixo": "2704",
        "ts_circular": True, "saida_ts": "23:00:00",
    })
    assert resp.status_code == 422, resp.text


# ============================================================================
# 10 — GET /turnos/ativo NÃO é capturado por /turnos/{turno_id}
# ============================================================================

def test_turnos_ativo_nao_capturado_por_turno_id(ambiente):
    _como(ambiente, "FISCAL")
    sem_turno = ambiente["http"].get("/fiscalizacao/turnos/ativo")
    assert sem_turno.status_code == 200, sem_turno.text
    assert sem_turno.json() is None  # nunca um 422 de "ativo não é um UUID válido"

    aberto = ambiente["http"].post("/fiscalizacao/turnos", json={
        "ponto_codigo": "PQ_TESTE", "periodo": "1", "linhas": ["1726-10"],
    })
    assert aberto.status_code == 201, aberto.text

    ativo = ambiente["http"].get("/fiscalizacao/turnos/ativo")
    assert ativo.status_code == 200, ativo.text
    assert ativo.json()["id"] == aberto.json()["id"]


# ============================================================================
# 11 — 🔴 Estado derivado (D7): ATRASADA/AGUARDANDO calculado na leitura,
# nada gravado entre as duas leituras
# ============================================================================

def test_estado_atrasada_aguardando_calculado_na_leitura(ambiente, monkeypatch):
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="1")
    _criar_partida_programada(
        ambiente, linha_codigo="1726-10", tipo_dia="UTIL", numero_tabela=1, sequencia=1,
        terminal="TP", horario=time(8, 0), periodo="1", vigencia=hoje,
    )
    _criar_partida_programada(
        ambiente, linha_codigo="1726-10", tipo_dia="UTIL", numero_tabela=2, sequencia=1,
        terminal="TP", horario=time(22, 0), periodo="1", vigencia=hoje,
    )

    _congelar_relogio(monkeypatch, datetime(2026, 8, 21, 10, 0, tzinfo=FUSO_OPERACAO))
    _como(ambiente, "FISCAL")

    resp = ambiente["http"].get(f"/fiscalizacao/turnos/{turno.id}/partidas")
    assert resp.status_code == 200, resp.text
    itens = resp.json()["1726-10"]
    estados = {item["numero_tabela"]: item["estado"] for item in itens}
    assert estados[1] == "ATRASADA"    # 08:00 já passou das 10:00
    assert estados[2] == "AGUARDANDO"  # 22:00 ainda não chegou

    with Session(ambiente["engine"]) as db:
        gravados = db.execute(select(RegistroPartida).where(RegistroPartida.turno_id == turno.id)).scalars().all()
    assert gravados == []  # 🔴 nada foi gravado pela leitura

    # Segunda leitura, relógio adiantado — mesma partida 22:00 vira ATRASADA.
    _congelar_relogio(monkeypatch, datetime(2026, 8, 21, 23, 0, tzinfo=FUSO_OPERACAO))
    resp2 = ambiente["http"].get(f"/fiscalizacao/turnos/{turno.id}/partidas")
    estados2 = {item["numero_tabela"]: item["estado"] for item in resp2.json()["1726-10"]}
    assert estados2[2] == "ATRASADA"


# ============================================================================
# 12 — 🔴 D6: 40 programadas, 39 realizadas, 1 viagem extra → os quatro
# números ao mesmo tempo
# ============================================================================

def test_d6_realizadas_programadas_extras_e_total_separados(ambiente):
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="2")

    for i in range(40):
        _criar_partida_programada(
            ambiente, linha_codigo="1726-10", tipo_dia="UTIL", numero_tabela=(i % 9) + 1, sequencia=i + 1,
            terminal="TP", horario=time(13, 0), periodo="2", vigencia=hoje,
        )
    for i in range(39):
        _criar_registro(
            ambiente, turno_id=turno.id, linha_codigo="1726-10", numero_tabela=(i % 9) + 1,
            terminal="TP", horario_programado=time((14 + i // 9) % 24, i % 60), resultado="REALIZADA",
        )
    _criar_evento(ambiente, turno_id=turno.id, linha_codigo="1726-10", tipo="VIAGEM_EXTRA")

    with Session(ambiente["engine"]) as db:
        turno_db = db.get(Turno, turno.id)
        agregado = calcular_fechamento_linha(db, turno_db, "1726-10")

    assert agregado["programadas"] == 40
    assert agregado["realizadas_programadas"] == 39
    assert agregado["extras"] == 1
    assert agregado["realizadas_total"] == 40
    assert agregado["perdidas"] == 1


# ============================================================================
# 13 — 🔴 D4: 1 perda por OUTRO + 1 evento avulso de RA → perdidas=1 e
# contador RA=1, sem que a R.A vire perda
# ============================================================================

def test_d4_evento_avulso_nao_vira_perda(ambiente):
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="1")

    _criar_partida_programada(
        ambiente, linha_codigo="1726-10", tipo_dia="UTIL", numero_tabela=1, sequencia=1,
        terminal="TP", horario=time(8, 0), periodo="1", vigencia=hoje,
    )
    _criar_registro(
        ambiente, turno_id=turno.id, linha_codigo="1726-10", numero_tabela=1, terminal="TP",
        horario_programado=time(8, 0), resultado="PERDIDA", motivo="OUTRO",
        motivo_outro="sem condobus",
    )
    # Evento avulso de R.A — aconteceu, mas não custou viagem nenhuma.
    _criar_evento(ambiente, turno_id=turno.id, linha_codigo="1726-10", tipo="RA")

    with Session(ambiente["engine"]) as db:
        turno_db = db.get(Turno, turno.id)
        agregado = calcular_fechamento_linha(db, turno_db, "1726-10")

    assert agregado["perdidas"] == 1  # programadas(1) - realizadas_programadas(0)
    assert agregado["contadores"]["RA"] == 1
    # A perda em si foi por OUTRO, não por RA — D4: motivo OUTRO nunca cria evento.
    with Session(ambiente["engine"]) as db:
        eventos = db.execute(select(EventoTurno).where(EventoTurno.turno_id == turno.id)).scalars().all()
    assert len(eventos) == 1
    assert eventos[0].tipo == "RA"
    assert eventos[0].registro_partida_id is None  # avulso, não vinculado a nenhuma partida


# ============================================================================
# 14 — POST /fechar funciona com partidas sem resposta (D3 — avisa, não bloqueia)
# ============================================================================

def test_fechar_turno_funciona_com_pendencias(ambiente):
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="1")
    _criar_partida_programada(
        ambiente, linha_codigo="1726-10", tipo_dia="UTIL", numero_tabela=1, sequencia=1,
        terminal="TP", horario=time(8, 0), periodo="1", vigencia=hoje,
    )
    _como(ambiente, "FISCAL")

    prontidao = ambiente["http"].get(f"/fiscalizacao/turnos/{turno.id}/prontidao")
    assert prontidao.status_code == 200, prontidao.text
    assert prontidao.json()["pronto"] is False
    tipos = {p["tipo"] for p in prontidao.json()["pendencias"]}
    assert "BAITA_FALTANDO" in tipos
    assert "ANTI_BAITA_FALTANDO" in tipos
    assert "PASTAS_NAO_INFORMADAS" in tipos

    fechar = ambiente["http"].post(f"/fiscalizacao/turnos/{turno.id}/fechar")
    assert fechar.status_code == 200, fechar.text
    assert fechar.json()["status"] == "FECHADO"


# ============================================================================
# 15 — Gerador: fechamento completo, caractere a caractere
# ============================================================================

def test_gerador_fechamento_texto_exato(ambiente):
    hoje = date(2026, 8, 21)
    turno = _criar_turno(
        ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="2", linhas=("2032-10",),
        refeicao_inicio=time(21, 55), refeicao_fim=time(22, 55), pastas_prefixo="2704",
    )

    for i in range(40):
        _criar_partida_programada(
            ambiente, linha_codigo="2032-10", tipo_dia="UTIL", numero_tabela=(i % 9) + 1, sequencia=i + 1,
            terminal="TP", horario=time(13, 0), periodo="2", vigencia=hoje,
        )
    for i in range(39):
        _criar_registro(
            ambiente, turno_id=turno.id, linha_codigo="2032-10", numero_tabela=(i % 9) + 1,
            terminal="TP", horario_programado=time((14 + i // 9) % 24, i % 60), resultado="REALIZADA",
        )
    _criar_registro(
        ambiente, turno_id=turno.id, linha_codigo="2032-10", numero_tabela=8, terminal="TP",
        horario_programado=time(20, 20), resultado="PERDIDA", motivo="OUTRO",
        motivo_outro="motorista 7277 está sem condobus",
    )
    _criar_evento(ambiente, turno_id=turno.id, linha_codigo="2032-10", tipo="RA")
    _criar_evento(ambiente, turno_id=turno.id, linha_codigo="2032-10", tipo="ATRASO_GARAGEM")

    with Session(ambiente["engine"]) as db:
        db.add(ObservacaoTurno(
            id=uuid4(), turno_id=turno.id, linha_codigo="2032-10",
            texto="Tabela 08, pegada de garagem 15:45 com saída prevista no TP 16:00 chegou com 10 minutos de atraso.",
        ))
        db.add(Baita(
            id=uuid4(), turno_id=turno.id, linha_codigo="2032-10", tipo="ANTI_BAITA",
            prefixo="2494", motorista_re="14045", cobrador_re="18935",
            saida_tp=time(23, 20), ts_circular=True,
        ))
        db.add(Baita(
            id=uuid4(), turno_id=turno.id, linha_codigo="2032-10", tipo="BAITA",
            prefixo="2704", motorista_re="8201", cobrador_re="19005",
            saida_tp=time(23, 45), ts_circular=True,
        ))
        db.commit()

        turno_db = db.get(Turno, turno.id)
        mensagens = montar_fechamento(db, turno_db.id)

    assert len(mensagens) == 1
    esperado = (
        "Sambaiba G3\n"
        "SEGUE ABAIXO FECHAMENTO... 21/08/2026\n"
        "LINHA: 2032-10\n"
        "PARTIDAS PROGRAMADAS: 2° período 40\n"
        "PARTIDAS REALIZADAS: 2° período 39\n"
        " FALTA DE OPERADORES: 00\n"
        " R.A: 01\n"
        " S.O.S: 00\n"
        " ATRASO DE GARAGEM 01\n"
        " TROCA OPERACIONAL 00\n"
        " VIAGEM EXTRA 00\n"
        "OBS:\n"
        " 👉🏼 FISCAL 70001, refeição das 21:55 às 22:55.\n"
        " 👉🏼 Tabela 08, pegada de garagem 15:45 com saída prevista no TP 16:00 chegou com 10 minutos de atraso.\n"
        " 👉🏼 Tabela 08, motorista 7277 está sem condobus, viagem das 20:20 não realizada.\n"
        "ANTI-BAITA\n"
        "CARRO: 2494\n"
        "MOT: 14045\n"
        "COB: 18935\n"
        "SAÍDA DO TP 23:20\n"
        "SAÍDA DO TS circular\n"
        "BAITA\n"
        "CARRO: 2704\n"
        "MOT: 8201\n"
        "COB: 19005\n"
        "SAÍDA DO TP 23:45\n"
        "SAÍDA DO TS circular\n"
        "PASTAS TP/TS\n"
        "CARRO: 2704\n"
        "FISCAL RE... 70001"
    )
    assert mensagens[0] == esperado


# ============================================================================
# 16 — Gerador: sem pastas_prefixo → bloco PASTAS TP/TS não aparece; nenhum
# None no texto
# ============================================================================

def test_gerador_sem_pastas_omite_bloco_e_sem_none(ambiente):
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="1", linhas=("1726-10",))
    # Turno sem BAITA/ANTI_BAITA cadastrado nenhum, sem refeição, sem pastas —
    # o cenário mais vazio possível, exatamente onde "None" vazaria se algum
    # campo esquecesse o fallback "—".

    with Session(ambiente["engine"]) as db:
        turno_db = db.get(Turno, turno.id)
        mensagens = montar_fechamento(db, turno_db.id)

    assert len(mensagens) == 1
    texto = mensagens[0]
    assert "PASTAS TP/TS" not in texto
    assert "None" not in texto
    assert "null" not in texto
    # Campos obrigatórios ausentes saem como "—", nunca em branco.
    assert "CARRO: —" in texto
    assert "MOT: —" in texto
    assert "COB: —" in texto
    assert "SAÍDA DO TP —" in texto
    assert "SAÍDA DO TS —" in texto


# ============================================================================
# Smoke — modelos/schemas criam tabela e o app registra as rotas (Bloco B)
# ============================================================================

def test_app_registra_rotas_fiscalizacao():
    """⚠️ Lê pelo OpenAPI, NÃO por app.routes. A partir de certa versão do
    FastAPI, cada include_router() vira UM objeto _IncludedRouter em
    app.routes e as rotas dos módulos deixam de aparecer ali — foi o que
    fez este teste falhar sem que nada estivesse quebrado (24/08/2026).
    O OpenAPI é a superfície pública e preserva a ordem de declaração."""
    caminhos = list(app.openapi()["paths"].keys())
    assert "/fiscalizacao/turnos/ativo" in caminhos
    assert "/fiscalizacao/turnos/{turno_id}" in caminhos

    # ⛔ ordem: rota literal SEMPRE antes da rota com path parameter, senão
    # "ativo" chega como se fosse um turno_id. Mesma classe de bug que já
    # apareceu quatro vezes neste projeto.
    assert caminhos.index("/fiscalizacao/turnos/ativo") < caminhos.index("/fiscalizacao/turnos/{turno_id}")

    # D39 — as duas rotas literais do painel, pela mesma razão: sem isto,
    # "ao-vivo" e "turnos" seriam capturados como código de linha.
    assert caminhos.index("/fiscalizacao/painel/ao-vivo") < caminhos.index("/fiscalizacao/painel/{linha_codigo}")
    assert caminhos.index("/fiscalizacao/painel/turnos") < caminhos.index("/fiscalizacao/painel/{linha_codigo}")


# ============================================================================
# Bloco C — importação da grade (parser + endpoint)
# ============================================================================

# ⛔ Sem caminho de máquina de ninguém aqui: o repositório é público. A
# planilha real da escala gerencial vive em _handoff-claude/, que está no
# .gitignore — ou seja, ela NÃO acompanha o clone. Quem tiver o arquivo roda
# estes quatro testes; quem não tiver, PULA. ⛔ Nunca falhar por ausência de
# um arquivo que o repositório não distribui.
_RAIZ_REPO = Path(__file__).resolve().parents[2]
_PASTA1_XLSX = Path(os.environ.get("FISCALIZACAO_PASTA1_XLSX") or (
    _RAIZ_REPO / "_handoff-claude" / "Coordenadoria_Sambaiba" / "analise-pre-juncao"
    / "Escalas Gerenciais" / "Pasta1.xlsx"
))

_exige_planilha_real = pytest.mark.skipif(
    not _PASTA1_XLSX.exists(),
    reason=(
        "planilha real da escala gerencial não encontrada — aponte "
        "FISCALIZACAO_PASTA1_XLSX para ela, ou deixe-a em _handoff-claude/"
    ),
)


@_exige_planilha_real
def test_parser_conta_66_tp_e_66_ts_no_arquivo_real():
    import openpyxl

    wb = openpyxl.load_workbook(_PASTA1_XLSX, data_only=True)
    ws = wb["1726-10"]
    partidas = _parse_aba(ws)

    tp = [p for p in partidas if p["terminal"] == "TP"]
    ts = [p for p in partidas if p["terminal"] == "TS"]
    # Referência conhecida (§7.3 do prompt, conferida contra o resumo da
    # própria planilha: VIAG. IDA 66 / VIAG. VOLTA 66).
    assert len(tp) == 66, f"esperado 66 partidas TP, veio {len(tp)}"
    assert len(ts) == 66, f"esperado 66 partidas TS, veio {len(ts)}"


@_exige_planilha_real
def test_parser_classifica_periodo_por_tabela():
    import openpyxl

    wb = openpyxl.load_workbook(_PASTA1_XLSX, data_only=True)
    ws = wb["1726-10"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, row in enumerate(rows) if row and str(row[0]).strip().upper() == "TABELAS")
    tabela_cols = {j: int(c) for j, c in enumerate(rows[header_idx]) if j and isinstance(c, (int, float)) and 1 <= c <= 30}
    limites = _parse_limites_periodo(rows, tabela_cols)

    # Tabela 1: TERM. 1º PERIODO = 13:25, INICIO 2º PERIODO = 13:35 (lidos
    # diretamente do arquivo real na sessão de inspeção).
    term1, inicio2 = limites[1]
    assert term1 == time(13, 25)
    assert inicio2 == time(13, 35)
    assert _classificar_periodo(time(4, 30), (term1, inicio2)) == "1"
    assert _classificar_periodo(time(23, 20), (term1, inicio2)) == "2"


@_exige_planilha_real
def test_upload_escala_endpoint_importa_arquivo_real(ambiente):
    _como(ambiente, "COORDENADOR")  # escrita_escala liberada
    with open(_PASTA1_XLSX, "rb") as arquivo:
        resp = ambiente["http"].post(
            "/fiscalizacao/escalas/upload",
            files={"file": ("Pasta1.xlsx", arquivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            params={"tipo_dia": "UTIL", "vigencia": "2026-08-22"},
        )
    assert resp.status_code == 200, resp.text
    corpo = resp.json()
    linha = next(item for item in corpo["linhas"] if item["linha"] == "1726-10")
    assert linha["partidas_importadas"] == 132  # 66 TP + 66 TS

    with Session(ambiente["engine"]) as db:
        total = db.execute(
            select(PartidaProgramada).where(PartidaProgramada.linha_codigo == "1726-10")
        ).scalars().all()
    assert len(total) == 132


@_exige_planilha_real
def test_upload_escala_fiscal_sem_permissao_nega_403(ambiente):
    _como(ambiente, "FISCAL")
    with open(_PASTA1_XLSX, "rb") as arquivo:
        resp = ambiente["http"].post(
            "/fiscalizacao/escalas/upload",
            files={"file": ("Pasta1.xlsx", arquivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 403, resp.text


# ─── upload de 50 MB → 413, sem carregar tudo em memória (mesmo padrão de
# test_pre_ocorrencia.py::test_upload_publico_50mb_retorna_413) ────────────────

def test_upload_escala_50mb_retorna_413(ambiente):
    _como(ambiente, "COORDENADOR")  # escrita_escala liberada
    conteudo_grande = b"x" * (50 * 1024 * 1024)
    resp = ambiente["http"].post(
        "/fiscalizacao/escalas/upload",
        files={"file": ("escala.xlsx", conteudo_grande, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 413, resp.text


# ============================================================================
# Bloco D — D34: registro fora da grade precisa aparecer na listagem
# ============================================================================

def test_partida_fora_da_grade_aparece_na_listagem(ambiente):
    """🔴 O teste que provava o defeito central do D34: antes da correção,
    _itens_partidas devolvia [] assim que não achava vigência, e o
    registro gravado ficava invisível. Turno sem NENHUMA partida_programada
    + uma anormalidade registrada → a listagem tem que devolver 1 item,
    com fora_da_grade=true e partida_programada_id nulo."""
    turno = _criar_turno(ambiente)  # sem nenhuma partida_programada
    _como(ambiente, "FISCAL")

    marcar = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        "linha_codigo": "1726-10", "terminal": "TP", "horario_programado": "08:00:00",
        "resultado": "PERDIDA", "motivo": "RA", "prefixo": "1721",
    })
    assert marcar.status_code == 200, marcar.text

    resp = ambiente["http"].get(f"/fiscalizacao/turnos/{turno.id}/partidas")
    assert resp.status_code == 200, resp.text
    itens = resp.json()["1726-10"]
    assert len(itens) == 1
    assert itens[0]["fora_da_grade"] is True
    assert itens[0]["partida_programada_id"] is None
    assert itens[0]["estado"] == "PERDIDA"


# ============================================================================
# Bloco D — D33: upsert sem tabela não duplica (guarda de NULL na UNIQUE)
# ============================================================================

def test_upsert_sem_tabela_duas_vezes_nao_duplica(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    payload_base = {"linha_codigo": "1726-10", "terminal": "TP", "horario_programado": "09:00:00"}

    primeira = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        **payload_base, "resultado": "PERDIDA", "motivo": "OUTRO", "motivo_outro": "sem condobus",
    })
    assert primeira.status_code == 200, primeira.text

    segunda = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        **payload_base, "resultado": "PERDIDA", "motivo": "OUTRO", "motivo_outro": "confirmado sem condobus",
    })
    assert segunda.status_code == 200, segunda.text
    assert segunda.json()["id"] == primeira.json()["id"]  # mesmo registro, não um novo

    with Session(ambiente["engine"]) as db:
        registros = db.execute(select(RegistroPartida).where(RegistroPartida.turno_id == turno.id)).scalars().all()
    assert len(registros) == 1
    assert registros[0].motivo_outro == "confirmado sem condobus"


# ============================================================================
# Bloco D — D35: contagem informada no fechamento, quando não há grade
# ============================================================================

def test_fechamento_sem_grade_usa_contagem_informada(ambiente):
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="1", linhas=("1726-10",))
    _como(ambiente, "FISCAL")

    patch = ambiente["http"].patch(f"/fiscalizacao/turnos/{turno.id}/linhas/1726-10", json={
        "programadas_informadas": 20, "realizadas_informadas": 18, "extras_informadas": 2,
    })
    assert patch.status_code == 200, patch.text

    with Session(ambiente["engine"]) as db:
        turno_db = db.get(Turno, turno.id)
        agregado = calcular_fechamento_linha(db, turno_db, "1726-10")
    assert agregado["fonte"] == "INFORMADO"
    assert agregado["programadas"] == 20
    assert agregado["realizadas_programadas"] == 18
    assert agregado["extras"] == 2
    assert agregado["realizadas_total"] == 20  # D6: 18 + 2
    assert agregado["perdidas"] == 2


def test_fechamento_sem_grade_sem_contagem_informada_sai_traco(ambiente):
    """D35 — nulo nunca vira zero: sem grade e sem contagem informada, o
    gerador imprime "—" nas duas primeiras linhas, nunca "00"."""
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="1", linhas=("1726-10",))

    with Session(ambiente["engine"]) as db:
        turno_db = db.get(Turno, turno.id)
        mensagens = montar_fechamento(db, turno_db.id)

    assert len(mensagens) == 1
    texto = mensagens[0]
    assert "PARTIDAS PROGRAMADAS: 1° período —" in texto
    assert "PARTIDAS REALIZADAS: 1° período —" in texto


def test_fechamento_com_grade_ignora_informado(ambiente):
    """D35 — quando existe grade vigente, ela manda: o que o fiscal tiver
    informado (por engano, ou de um turno anterior) é ignorado."""
    hoje = date(2026, 8, 21)
    turno = _criar_turno(ambiente, data_referencia=hoje, tipo_dia="UTIL", periodo="1", linhas=("1726-10",))
    _criar_partida_programada(
        ambiente, linha_codigo="1726-10", tipo_dia="UTIL", numero_tabela=1, sequencia=1,
        terminal="TP", horario=time(8, 0), periodo="1", vigencia=hoje,
    )
    _como(ambiente, "FISCAL")

    patch = ambiente["http"].patch(f"/fiscalizacao/turnos/{turno.id}/linhas/1726-10", json={
        "programadas_informadas": 999, "realizadas_informadas": 999, "extras_informadas": 999,
    })
    assert patch.status_code == 200, patch.text

    with Session(ambiente["engine"]) as db:
        turno_db = db.get(Turno, turno.id)
        agregado = calcular_fechamento_linha(db, turno_db, "1726-10")
    assert agregado["fonte"] == "GRADE"
    assert agregado["programadas"] == 1  # da grade, não 999


# ============================================================================
# Bloco D — D36: prontidão sem grade cobra contagem, não partida sem resposta
# ============================================================================

def test_prontidao_sem_grade_cobra_contagem_nao_partida(ambiente):
    turno = _criar_turno(ambiente)  # sem partida_programada nenhuma
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].get(f"/fiscalizacao/turnos/{turno.id}/prontidao")
    assert resp.status_code == 200, resp.text
    tipos = {p["tipo"] for p in resp.json()["pendencias"]}
    assert "CONTAGEM_NAO_INFORMADA" in tipos
    assert "PARTIDAS_SEM_RESPOSTA" not in tipos


def test_prontidao_sem_refeicao_cobra(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    resp = ambiente["http"].get(f"/fiscalizacao/turnos/{turno.id}/prontidao")
    assert resp.status_code == 200, resp.text
    tipos = {p["tipo"] for p in resp.json()["pendencias"]}
    assert "REFEICAO_NAO_INFORMADA" in tipos


# ============================================================================
# Bloco D — D37: cadastro de pontos pela tela
# ============================================================================

def test_post_ponto_codigo_repetido_nega_409(ambiente):
    _como(ambiente, "FISCAL")
    primeiro = ambiente["http"].post("/fiscalizacao/pontos", json={
        "codigo": "PQ_NOVO", "nome": "Ponto Novo", "terminal": "TP", "linhas": ["1726-10"],
    })
    assert primeiro.status_code == 201, primeiro.text

    segundo = ambiente["http"].post("/fiscalizacao/pontos", json={
        "codigo": "PQ_NOVO", "nome": "Outro nome", "terminal": "TS", "linhas": ["9999-10"],
    })
    assert segundo.status_code == 409, segundo.text


# ============================================================================
# Bloco D — D38: coordenador atribui e remove as próprias linhas
# ============================================================================

def test_post_minha_linha_ja_atribuida_nega_409_sem_nome(ambiente):
    _como(ambiente, "COORDENADOR")
    primeiro = ambiente["http"].post("/fiscalizacao/minhas-linhas", json={
        "linha_codigo": "1726-10", "periodo": "1",
    })
    assert primeiro.status_code == 201, primeiro.text

    _como(ambiente, "COORDENADOR", usuario=_COORDENADOR_B)
    segundo = ambiente["http"].post("/fiscalizacao/minhas-linhas", json={
        "linha_codigo": "1726-10", "periodo": "1",
    })
    assert segundo.status_code == 409, segundo.text
    # ⛔ D38 — a mensagem de recusa não diz de quem é a linha.
    assert _COORDENADOR.nome not in segundo.text
    assert _COORDENADOR.re not in segundo.text


def test_delete_minha_linha_de_outro_nega_403(ambiente):
    _como(ambiente, "COORDENADOR")
    criar = ambiente["http"].post("/fiscalizacao/minhas-linhas", json={
        "linha_codigo": "1726-10", "periodo": "1",
    })
    assert criar.status_code == 201, criar.text

    _como(ambiente, "COORDENADOR", usuario=_COORDENADOR_B)
    remover = ambiente["http"].delete("/fiscalizacao/minhas-linhas/1726-10", params={"periodo": "1"})
    assert remover.status_code == 403, remover.text


# ============================================================================
# Bloco D — D35: a costura nova (PATCH contagem) exige escrita em
# fiscalizacao, não em fiscalizacao_painel
# ============================================================================

def test_coordenador_sem_escrita_fiscalizacao_no_patch_linha_nega_403(ambiente):
    turno = _criar_turno(ambiente)
    _como(ambiente, "COORDENADOR")
    resp = ambiente["http"].patch(f"/fiscalizacao/turnos/{turno.id}/linhas/1726-10", json={
        "programadas_informadas": 10,
    })
    assert resp.status_code == 403, resp.text


# ============================================================================
# Bloco D — D39: o painel ao vivo. Estes dois endpoints são o topo da tela do
# coordenador e foram a última costura do bloco — sem eles o painel chamava
# rota inexistente e as duas primeiras seções morriam em 404.
# ============================================================================

def test_painel_ao_vivo_lista_registro_do_fiscal_na_linha_do_coordenador(ambiente):
    """Ponta a ponta do D39: o fiscal marca uma perda; o coordenador que
    coordena aquela linha vê o registro no ao vivo, com o motivo, o RE de
    quem marcou e a marcação de que custou viagem (D4)."""
    _como(ambiente, "COORDENADOR")
    atribuir = ambiente["http"].post("/fiscalizacao/minhas-linhas", json={
        "linha_codigo": "1726-10", "periodo": "1",
    })
    assert atribuir.status_code == 201, atribuir.text

    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    marcar = ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        "linha_codigo": "1726-10", "terminal": "TP", "horario_programado": "18:20:00",
        "resultado": "PERDIDA", "motivo": "RA", "prefixo": "1721",
    })
    assert marcar.status_code == 200, marcar.text

    _como(ambiente, "COORDENADOR")
    resp = ambiente["http"].get("/fiscalizacao/painel/ao-vivo")
    assert resp.status_code == 200, resp.text
    itens = resp.json()
    assert len(itens) == 1, itens
    assert itens[0]["linha_codigo"] == "1726-10"
    assert itens[0]["tipo"] == "RA"
    assert itens[0]["custou_viagem"] is True
    assert itens[0]["fiscal_re"] == _FISCAL_A.re
    assert itens[0]["ponto_codigo"] == "PQ_TESTE"
    assert itens[0]["minutos_atras"] >= 0


def test_painel_ao_vivo_nao_conta_duas_vezes_o_evento_vinculado(ambiente):
    """D4 na tela: marcar PERDIDA com motivo que tem contador cria um
    evento_turno vinculado. O ao vivo lista o registro e IGNORA esse evento
    — senão a mesma perda apareceria duas vezes para quem olha o painel."""
    _como(ambiente, "COORDENADOR")
    ambiente["http"].post("/fiscalizacao/minhas-linhas", json={
        "linha_codigo": "1726-10", "periodo": "1",
    })
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        "linha_codigo": "1726-10", "terminal": "TP", "horario_programado": "19:10:00",
        "resultado": "PERDIDA", "motivo": "SOS", "prefixo": "1730",
    })

    with Session(ambiente["engine"]) as db:
        eventos = db.execute(
            select(EventoTurno).where(EventoTurno.turno_id == turno.id)
        ).scalars().all()
    assert len(eventos) == 1  # o vinculado existe...

    _como(ambiente, "COORDENADOR")
    itens = ambiente["http"].get("/fiscalizacao/painel/ao-vivo").json()
    assert len(itens) == 1  # ...e mesmo assim o ao vivo mostra UMA linha


def test_painel_ao_vivo_ignora_linha_que_nao_e_do_coordenador(ambiente):
    """O painel é das linhas de quem está logado (D39). Sem atribuição
    nenhuma (D40), a lista vem vazia — e isso é caso normal, não erro."""
    turno = _criar_turno(ambiente)
    _como(ambiente, "FISCAL")
    ambiente["http"].put(f"/fiscalizacao/turnos/{turno.id}/partidas", json={
        "linha_codigo": "1726-10", "terminal": "TP", "horario_programado": "20:00:00",
        "resultado": "PERDIDA", "motivo": "RA", "prefixo": "1721",
    })

    _como(ambiente, "COORDENADOR")  # nenhuma linha atribuída
    resp = ambiente["http"].get("/fiscalizacao/painel/ao-vivo")
    assert resp.status_code == 200, resp.text
    assert resp.json() == []


def test_painel_turnos_mostra_quem_esta_na_rua(ambiente):
    """D39 — turnos abertos nas minhas linhas, com o nome do fiscal e há
    quanto tempo ele não registra nada. Turno recém-aberto sem registro
    nenhum vem com minutos_sem_registrar nulo, que é diferente de zero."""
    _como(ambiente, "COORDENADOR")
    ambiente["http"].post("/fiscalizacao/minhas-linhas", json={
        "linha_codigo": "1726-10", "periodo": "1",
    })
    turno = _criar_turno(ambiente)

    resp = ambiente["http"].get("/fiscalizacao/painel/turnos")
    assert resp.status_code == 200, resp.text
    itens = resp.json()
    assert len(itens) == 1, itens
    assert itens[0]["fiscal_re"] == _FISCAL_A.re
    assert itens[0]["fiscal_nome"] == _FISCAL_A.nome
    assert itens[0]["ponto_codigo"] == "PQ_TESTE"
    assert itens[0]["linhas"] == ["1726-10"]
    assert itens[0]["minutos_sem_registrar"] is None


def test_fiscal_sem_painel_no_ao_vivo_nega_403(ambiente):
    """RBAC do módulo: FISCAL registra, não vê agregado (mesma decisão do
    controlador de acesso na Portaria). Vale para os dois endpoints novos."""
    _como(ambiente, "FISCAL")
    assert ambiente["http"].get("/fiscalizacao/painel/ao-vivo").status_code == 403
    assert ambiente["http"].get("/fiscalizacao/painel/turnos").status_code == 403
