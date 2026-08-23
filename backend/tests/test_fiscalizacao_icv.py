"""Módulo Fiscalização — Bloco E: ICV, bacia e painel do coordenador.

Testes escritos JUNTO de cada bloco (não represados para o fim — ver
_handoff-claude/PROMPT-fiscalizacao-bloco-E-icv.md, instrução extra do
Alisson). Mesmo padrão de test_fiscalizacao.py: SQLite em memória, sem
conftest.py, sqlite3.register_adapter para UUID.

⛔ Nenhum dado pessoal real — RE, prefixo e nome fictícios. Os números de
ICV usados nos testes são FICTÍCIOS/didáticos — não temos o arquivo real de
planilha à disposição (§10 do prompt) — escolhidos só para provar a
matemática (ponderação D22, prioridade D23, suspeita D25), não para
reproduzir uma garagem real.

⚠️ Assim como o resto do módulo (ver RegistroPartida em models/
fiscalizacao.py), os CHECK constraints deste bloco (lote IN ('E2','AR2'),
programadas >= 0, faixa_hora BETWEEN 0 AND 23) existem SÓ na migration SQL
— nenhum model do projeto declara CheckConstraint no SQLAlchemy. Por isso
os testes aqui não tentam provar CHECK via INSERT direto em SQLite (não
seria enforced nessa camada); a validação de entrada da API é coberta nos
schemas Pydantic (ver test_acao_coordenacao_* nos blocos seguintes, que
testam via endpoint HTTP, não via Session direta).
"""
import io
import sqlite3
import typing
import uuid as _uuid_mod
from datetime import date, time, timedelta
from uuid import uuid4

import openpyxl
import pytest
from fastapi import HTTPException, status
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, event, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.core.database import Base, get_db
from app.main import app
from app.models.cadastro import Funcionario
from app.models.fiscalizacao import (
    AcaoCoordenacao, Bacia, BaciaLinha, EventoTurno, IcvApurado, PartidaProgramada,
    Ponto, RegistroPartida, Turno,
)
from app.routers import fiscalizacao as fiscalizacao_router_mod
from app.services.icv import (
    _tipo_dia, calcular_icv_bacia_dia, calcular_icv_linha_dia, montar_placar_linha,
    ranking_prioridade,
)

sqlite3.register_adapter(_uuid_mod.UUID, lambda u: u.hex)

_TABELAS = [
    Funcionario.__table__, Bacia.__table__, BaciaLinha.__table__,
    IcvApurado.__table__, AcaoCoordenacao.__table__,
    Ponto.__table__, Turno.__table__, PartidaProgramada.__table__,
    RegistroPartida.__table__, EventoTurno.__table__,
]


@pytest.fixture
def db():
    engine = create_engine(
        "sqlite:///:memory:", poolclass=StaticPool, connect_args={"check_same_thread": False}
    )

    @event.listens_for(engine, "connect")
    def _attach(dbapi_conn, _):
        dbapi_conn.execute("ATTACH DATABASE ':memory:' AS fiscalizacao")
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine, tables=_TABELAS)
    with Session(engine) as session:
        yield session


# ============================================================================
# BLOCO 1 — migration 030: tabelas, defaults e constraints declaradas no
# SQLAlchemy (UniqueConstraint — os CHECK ficam só na migration SQL)
# ============================================================================

def test_bacia_meta_icv_default_98(db):
    """D29 — meta_icv tem DEFAULT 98.00 no schema, não hardcoded em cada
    INSERT do código de aplicação."""
    b = Bacia(codigo="BACIA_TESTE", nome="Bacia Teste")
    db.add(b)
    db.commit()
    db.refresh(b)
    assert float(b.meta_icv) == 98.00


def test_bacia_linha_unique_por_linha_vigencia_inicio(db):
    """A UNIQUE real (linha_codigo, vigencia_inicio) — não deixa duas
    vigências abrirem no mesmo dia para a mesma linha (ver "DESVIO" no
    cabeçalho da migration 030: não é UNIQUE(bacia_codigo, linha_codigo))."""
    db.add(Bacia(codigo="B1", nome="Bacia 1"))
    db.commit()
    db.add(BaciaLinha(bacia_codigo="B1", linha_codigo="0000-00", vigencia_inicio=date(2026, 8, 1)))
    db.commit()

    db.add(BaciaLinha(bacia_codigo="B1", linha_codigo="0000-00", vigencia_inicio=date(2026, 8, 1)))
    with pytest.raises(IntegrityError):
        db.commit()
    db.rollback()


def test_bacia_linha_mesma_linha_troca_de_bacia_com_vigencia(db):
    """D21 — a mesma linha pode aparecer em DUAS bacias diferentes, em
    vigências que não se sobrepõem (o caso real: 1726-10 mudou de bacia
    entre 26/05 e 19-21/08/2026). O schema tem que aceitar isso sem erro."""
    db.add(Bacia(codigo="B1", nome="Bacia 1"))
    db.add(Bacia(codigo="B2", nome="Bacia 2"))
    db.commit()
    db.add(BaciaLinha(
        bacia_codigo="B1", linha_codigo="0000-00",
        vigencia_inicio=date(2026, 5, 26), vigencia_fim=date(2026, 8, 18),
    ))
    db.add(BaciaLinha(
        bacia_codigo="B2", linha_codigo="0000-00",
        vigencia_inicio=date(2026, 8, 19), vigencia_fim=None,
    ))
    db.commit()  # não pode levantar

    linhas = db.query(BaciaLinha).filter_by(linha_codigo="0000-00").all()
    assert len(linhas) == 2
    assert {l.bacia_codigo for l in linhas} == {"B1", "B2"}


def test_icv_apurado_unique_linha_data_referencia(db):
    """A UNIQUE que o endpoint de upload usa como alvo de upsert (D25/§5
    do prompt) — reimportar o mesmo dia tem que ATUALIZAR, não duplicar;
    aqui provamos que a constraint em si impede a duplicata."""
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="1111-11", data_referencia=date(2026, 8, 19),
        programadas=100, realizadas_tp_ts=50, realizadas_ts_tp=45,
    ))
    db.commit()

    db.add(IcvApurado(
        id=uuid4(), linha_codigo="1111-11", data_referencia=date(2026, 8, 19),
        programadas=100, realizadas_tp_ts=48, realizadas_ts_tp=47,
    ))
    with pytest.raises(IntegrityError):
        db.commit()
    db.rollback()


def test_icv_apurado_sem_teto_realizadas_acima_de_programadas(db):
    """D30 — ICV pode passar de 100%: o banco não impede realizadas somadas
    maiores que programadas. Quem decide teto é a apresentação (services/
    icv.py), nunca o schema."""
    registro = IcvApurado(
        id=uuid4(), linha_codigo="2222-22", data_referencia=date(2026, 8, 19),
        programadas=20, realizadas_tp_ts=22, realizadas_ts_tp=0,
    )
    db.add(registro)
    db.commit()  # não deve levantar
    db.refresh(registro)
    assert registro.realizadas_tp_ts + registro.realizadas_ts_tp == 22 > registro.programadas


def test_icv_apurado_origem_default_planilha(db):
    registro = IcvApurado(
        id=uuid4(), linha_codigo="3333-33", data_referencia=date(2026, 8, 19), programadas=10,
    )
    db.add(registro)
    db.commit()
    db.refresh(registro)
    assert registro.origem == "PLANILHA"
    assert registro.suspeito is False


def test_acao_coordenacao_grava_com_funcionario_real(db):
    """D26 — ação da coordenação, ligada a um funcionário real (FK)."""
    func = Funcionario(id=uuid4(), re="70010", nome="Coordenador Teste", status="ATIVO")
    db.add(func)
    db.commit()

    acao = AcaoCoordenacao(
        id=uuid4(), linha_codigo="1726-10", data_referencia=date(2026, 8, 19),
        faixa_hora=18, descricao="Equipe reforçou o TP às 18h.",
        resultado_observado="ICV subiu de 94,62% para 99,46% em dois dias.",
        registrado_por=func.id,
    )
    db.add(acao)
    db.commit()
    db.refresh(acao)
    assert acao.registrado_por == func.id
    assert acao.faixa_hora == 18


# ============================================================================
# Fixture HTTP — mesmo padrão de test_fiscalizacao.py: dependências de
# permissão sobrescritas (vw_acesso_efetivo é view Postgres, não existe em
# SQLite), TestClient sobre app real.
# ============================================================================

def _dependency_de(annotated_type):
    for arg in typing.get_args(annotated_type)[1:]:
        dependency = getattr(arg, "dependency", None)
        if dependency is not None:
            return dependency
    raise RuntimeError("Depends não encontrado no tipo anotado")


_FISCAL = Funcionario(id=uuid4(), re="70020", nome="Fiscal ICV Teste")
_COORDENADOR = Funcionario(id=uuid4(), re="70021", nome="Coordenador ICV Teste")


@pytest.fixture
def ambiente():
    engine = create_engine(
        "sqlite:///:memory:", poolclass=StaticPool, connect_args={"check_same_thread": False}
    )

    @event.listens_for(engine, "connect")
    def _attach(dbapi_conn, _):
        dbapi_conn.execute("ATTACH DATABASE ':memory:' AS fiscalizacao")
        dbapi_conn.execute("PRAGMA foreign_keys=ON")

    Base.metadata.create_all(engine, tables=_TABELAS)
    with Session(engine) as setup:
        setup.add(Funcionario(id=_FISCAL.id, re=_FISCAL.re, nome=_FISCAL.nome, status="ATIVO"))
        setup.add(Funcionario(id=_COORDENADOR.id, re=_COORDENADOR.re, nome=_COORDENADOR.nome, status="ATIVO"))
        setup.commit()

    def _get_db_teste():
        session = Session(engine)
        try:
            yield session
        finally:
            session.close()

    dep_leitura_painel = _dependency_de(fiscalizacao_router_mod.LeituraPainel)
    dep_escrita_painel = _dependency_de(fiscalizacao_router_mod.EscritaPainel)

    app.dependency_overrides[get_db] = _get_db_teste

    yield {"engine": engine, "http": TestClient(app), "leitura_painel": dep_leitura_painel, "escrita_painel": dep_escrita_painel}

    app.dependency_overrides.pop(dep_leitura_painel, None)
    app.dependency_overrides.pop(dep_escrita_painel, None)
    app.dependency_overrides.pop(get_db, None)


def _negar():
    def _dep():
        raise HTTPException(status.HTTP_403_FORBIDDEN, detail="Sem permissão (fake, teste)")
    return _dep


def _permitir(usuario):
    def _dep():
        return usuario
    return _dep


def _como_fiscal(ambiente):
    app.dependency_overrides[ambiente["leitura_painel"]] = _negar()
    app.dependency_overrides[ambiente["escrita_painel"]] = _negar()


def _como_coordenador(ambiente):
    app.dependency_overrides[ambiente["leitura_painel"]] = _permitir(_COORDENADOR)
    app.dependency_overrides[ambiente["escrita_painel"]] = _permitir(_COORDENADOR)


# ============================================================================
# Fixture da planilha sintética — reproduz o layout de §5 do prompt.
# ⚠️ Dados FICTÍCIOS — não temos o arquivo real (§10 do prompt); falta
# validar contra um .xlsx real (ver PROGRESSO-2026-08-23.md).
# ============================================================================

_CABECALHO_ICV = ["aqa", "BACIA", "LOTE", "VIAGENS PROG.", "VIAGENS REAL TP-TS", "VIAGENS REAL TS-TP", "Total (% ICV)"]


def _construir_planilha_icv(data_str: str, linhas: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ICV GARAGEM 3"])
    ws.append([f"DATA: {data_str}"])
    ws.append(_CABECALHO_ICV)
    for linha in linhas:
        ws.append([
            linha["linha_codigo"],
            linha.get("bacia", ""),
            linha.get("lote", "E2"),
            linha["programadas"],
            linha.get("tp_ts", 0),
            linha.get("ts_tp", 0),
            linha.get("percentual"),
        ])
    total_prog = sum(l["programadas"] for l in linhas)
    ws.append(["TOTAL", "", "", total_prog, "", "", ""])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _planilha_sem_viagens_prog(data_str: str) -> bytes:
    """Formato de maio (D22) — sem coluna de contagem, só percentual."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ICV GARAGEM 3"])
    ws.append([f"DATA: {data_str}"])
    ws.append(["LINHA", "BACIA", "LOTE", "3 (% ICV)", "Media (% ICV)", "Total (% ICV)"])
    ws.append(["0000-00", "BACIA X", "E2", 95.0, 96.0, 95.5])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ============================================================================
# BLOCO 2 — importador da planilha de ICV (§5) + endpoint de upload
# ============================================================================

def test_upload_icv_fiscal_nega_403(ambiente):
    _como_fiscal(ambiente)
    conteudo = _construir_planilha_icv("19/08/2026", [
        {"linha_codigo": "0000-00", "programadas": 100, "tp_ts": 50, "ts_tp": 45, "percentual": 95.0},
    ])
    resp = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv.xlsx", conteudo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 403, resp.text


def test_upload_icv_coordenador_importa(ambiente):
    _como_coordenador(ambiente)
    conteudo = _construir_planilha_icv("19/08/2026", [
        {"linha_codigo": "0000-00", "bacia": "BACIA TESTE", "programadas": 100, "tp_ts": 50, "ts_tp": 45, "percentual": 95.0},
    ])
    resp = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv.xlsx", conteudo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    corpo = resp.json()
    assert corpo["linhas_lidas"] == 1
    assert corpo["linhas_gravadas"] == 1
    assert corpo["data_referencia"] == "2026-08-19"

    with Session(ambiente["engine"]) as db:
        registro = db.execute(select(IcvApurado).where(IcvApurado.linha_codigo == "0000-00")).scalar_one()
        assert registro.programadas == 100
        assert registro.realizadas_tp_ts == 50
        assert registro.realizadas_ts_tp == 45
        # D21 — a coluna BACIA alimenta bacia_linha da vigência.
        vinculo = db.execute(select(BaciaLinha).where(BaciaLinha.linha_codigo == "0000-00")).scalar_one()
        assert vinculo.vigencia_inicio == date(2026, 8, 19)
        assert vinculo.vigencia_fim is None


def test_upload_icv_formato_de_maio_recusado(ambiente):
    """D22 — sem VIAGENS PROG., o arquivo inteiro é recusado (400), nunca
    importado parcialmente."""
    _como_coordenador(ambiente)
    conteudo = _planilha_sem_viagens_prog("26/05/2026")
    resp = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv-maio.xlsx", conteudo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400, resp.text
    with Session(ambiente["engine"]) as db:
        assert db.execute(select(IcvApurado)).scalars().all() == []


def test_upload_icv_reimportar_mesmo_dia_atualiza_nao_duplica(ambiente):
    _como_coordenador(ambiente)
    primeira = _construir_planilha_icv("19/08/2026", [
        {"linha_codigo": "0000-00", "programadas": 100, "tp_ts": 50, "ts_tp": 45, "percentual": 95.0},
    ])
    resp1 = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv.xlsx", primeira, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp1.status_code == 200, resp1.text
    assert resp1.json()["linhas_gravadas"] == 1

    segunda = _construir_planilha_icv("19/08/2026", [
        {"linha_codigo": "0000-00", "programadas": 100, "tp_ts": 60, "ts_tp": 38, "percentual": 98.0},
    ])
    resp2 = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv2.xlsx", segunda, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp2.status_code == 200, resp2.text
    assert resp2.json()["linhas_atualizadas"] == 1
    assert resp2.json()["linhas_gravadas"] == 0

    with Session(ambiente["engine"]) as db:
        registros = db.execute(select(IcvApurado).where(IcvApurado.linha_codigo == "0000-00")).scalars().all()
    assert len(registros) == 1
    assert registros[0].realizadas_tp_ts == 60


def test_upload_icv_suspeita_dois_contadores_iguais_e_icv_nao_100(ambiente):
    """D25 — dois dias com os DOIS contadores idênticos e ICV < 100% marca
    o segundo como suspeito."""
    _como_coordenador(ambiente)
    dia19 = _construir_planilha_icv("19/08/2026", [
        {"linha_codigo": "1773-10", "programadas": 120, "tp_ts": 56, "ts_tp": 53, "percentual": 90.83},
    ])
    r1 = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv19.xlsx", dia19, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r1.status_code == 200, r1.text
    assert r1.json()["suspeitas"] == []

    dia20 = _construir_planilha_icv("20/08/2026", [
        {"linha_codigo": "1773-10", "programadas": 120, "tp_ts": 56, "ts_tp": 53, "percentual": 90.83},
    ])
    r2 = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv20.xlsx", dia20, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r2.status_code == 200, r2.text
    assert len(r2.json()["suspeitas"]) == 1
    assert r2.json()["suspeitas"][0]["linha_codigo"] == "1773-10"

    with Session(ambiente["engine"]) as db:
        registro = db.execute(
            select(IcvApurado).where(IcvApurado.linha_codigo == "1773-10", IcvApurado.data_referencia == date(2026, 8, 20))
        ).scalar_one()
        assert registro.suspeito is True
        assert registro.suspeito_motivo is not None


def test_upload_icv_falso_positivo_linha_100_por_cento_nao_e_suspeita(ambiente):
    """D25 — o teste que impede a detecção de virar ruído: linha a 100%
    repetindo os mesmos contadores NÃO é suspeita."""
    _como_coordenador(ambiente)
    dia19 = _construir_planilha_icv("19/08/2026", [
        {"linha_codigo": "1721-10", "programadas": 60, "tp_ts": 30, "ts_tp": 30, "percentual": 100.0},
    ])
    ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv19.xlsx", dia19, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    dia20 = _construir_planilha_icv("20/08/2026", [
        {"linha_codigo": "1721-10", "programadas": 60, "tp_ts": 30, "ts_tp": 30, "percentual": 100.0},
    ])
    r2 = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv20.xlsx", dia20, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r2.status_code == 200, r2.text
    assert r2.json()["suspeitas"] == []
    with Session(ambiente["engine"]) as db:
        registro = db.execute(
            select(IcvApurado).where(IcvApurado.linha_codigo == "1721-10", IcvApurado.data_referencia == date(2026, 8, 20))
        ).scalar_one()
        assert registro.suspeito is False


def test_upload_icv_ts_tp_vazio_linha_circular_vira_zero(ambiente):
    """VIAGENS REAL TS-TP vazio é normal (linha circular, §5) → grava 0,
    não erro."""
    _como_coordenador(ambiente)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ICV GARAGEM 3"])
    ws.append(["DATA: 19/08/2026"])
    ws.append(_CABECALHO_ICV)
    ws.append(["2024-10", "BACIA X", "E2", 50, 50, None, 100.0])  # circular — TS-TP vazio
    ws.append(["TOTAL", "", "", 50, "", "", ""])
    buffer = io.BytesIO()
    wb.save(buffer)

    resp = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["erros"] == []
    with Session(ambiente["engine"]) as db:
        registro = db.execute(select(IcvApurado).where(IcvApurado.linha_codigo == "2024-10")).scalar_one()
        assert registro.realizadas_ts_tp == 0


def test_upload_icv_codigo_ilegivel_reportado(ambiente):
    """Código corrompido pelo Excel em notação científica — importa como
    texto e reporta para conferência humana, não descarta."""
    _como_coordenador(ambiente)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ICV GARAGEM 3"])
    ws.append(["DATA: 19/08/2026"])
    ws.append(_CABECALHO_ICV)
    ws.append([2.13e-08, "BACIA X", "E2", 40, 20, 19, 97.5])
    ws.append(["TOTAL", "", "", 40, "", "", ""])
    buffer = io.BytesIO()
    wb.save(buffer)

    resp = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    corpo = resp.json()
    assert len(corpo["codigos_ilegiveis"]) == 1
    with Session(ambiente["engine"]) as db:
        registros = db.execute(select(IcvApurado)).scalars().all()
    assert len(registros) == 1  # não descartou a linha


def test_upload_icv_divergencia_percentual_reportada(ambiente):
    """D28 — recalcula o percentual e compara com a coluna da planilha;
    divergência além de 0,01 é reportada, não escondida."""
    _como_coordenador(ambiente)
    conteudo = _construir_planilha_icv("19/08/2026", [
        # (50+45)/100 = 95.0%, mas a planilha diz 90.0% — divergência.
        {"linha_codigo": "0000-00", "programadas": 100, "tp_ts": 50, "ts_tp": 45, "percentual": 90.0},
    ])
    resp = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv.xlsx", conteudo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200, resp.text
    divergentes = resp.json()["divergentes_percentual"]
    assert len(divergentes) == 1
    assert divergentes[0]["linha_codigo"] == "0000-00"
    assert divergentes[0]["icv_calculado"] == 95.0


def test_upload_icv_50mb_retorna_413(ambiente):
    _como_coordenador(ambiente)
    conteudo_grande = b"x" * (50 * 1024 * 1024)
    resp = ambiente["http"].post(
        "/fiscalizacao/icv/upload",
        files={"file": ("icv.xlsx", conteudo_grande, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 413, resp.text


# ============================================================================
# BLOCO 3 — app/services/icv.py: ICV ponderado (D22), prioridade por perda
# absoluta (D23), duas fontes (D20), divergência de denominador (D28),
# acima de 100% (D30). Números FICTÍCIOS — não temos a planilha real (§10
# do prompt); escolhidos só para provar a matemática.
# ============================================================================

def test_d22_bacia_ponderado_nao_e_media_simples(db):
    """A prova de que a média simples não voltou por acidente: uma linha
    grande (248 programadas) perto de 89% e uma pequena (13 programadas)
    perto de 92% — a média simples dos dois percentuais esconde o volume
    da linha grande."""
    hoje = date(2026, 8, 19)
    db.add(Bacia(codigo="B1", nome="Bacia Teste"))
    db.commit()
    db.add(BaciaLinha(bacia_codigo="B1", linha_codigo="LINHA_A", vigencia_inicio=date(2026, 1, 1)))
    db.add(BaciaLinha(bacia_codigo="B1", linha_codigo="LINHA_B", vigencia_inicio=date(2026, 1, 1)))
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="LINHA_A", data_referencia=hoje,
        programadas=248, realizadas_tp_ts=110, realizadas_ts_tp=110,  # 220/248 = 88,71%
    ))
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="LINHA_B", data_referencia=hoje,
        programadas=13, realizadas_tp_ts=6, realizadas_ts_tp=6,  # 12/13 = 92,31%
    ))
    db.commit()

    resultado = calcular_icv_bacia_dia(db, "B1", hoje)

    icv_a, icv_b = 220 / 248 * 100, 12 / 13 * 100
    media_simples = round((icv_a + icv_b) / 2, 2)

    assert resultado["programadas"] == 261
    assert resultado["realizadas"] == 232
    ponderado_esperado = round(232 / 261 * 100, 2)
    assert resultado["icv_ponderado"] == ponderado_esperado
    assert resultado["icv_ponderado"] != media_simples  # 🔴 este assert é o ponto do teste
    assert resultado["meta_icv"] == 98.00


def test_d23_ranking_ordena_por_perda_absoluta_nao_percentual(db):
    """Linha A (212 programadas, percentual MELHOR) aparece ACIMA da linha
    B (13 programadas, percentual pior) no ranking — porque perde mais
    viagem em absoluto, apesar do percentual melhor."""
    hoje = date(2026, 8, 19)
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="LINHA_A", data_referencia=hoje,
        programadas=212, realizadas_tp_ts=101, realizadas_ts_tp=101,  # 202/212 = 95,28% — perda 10
    ))
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="LINHA_B", data_referencia=hoje,
        programadas=13, realizadas_tp_ts=6, realizadas_ts_tp=6,  # 12/13 = 92,31% — perda 1
    ))
    db.commit()

    resultado = ranking_prioridade(db, hoje)
    codigos = [r["linha_codigo"] for r in resultado]
    assert codigos.index("LINHA_A") < codigos.index("LINHA_B")

    item_a = next(r for r in resultado if r["linha_codigo"] == "LINHA_A")
    item_b = next(r for r in resultado if r["linha_codigo"] == "LINHA_B")
    assert item_a["icv_oficial"] > item_b["icv_oficial"]       # A tem percentual MELHOR
    assert item_a["perda_absoluta"] > item_b["perda_absoluta"]  # mas perde mais em absoluto


def test_d28_divergencia_denominador_nenhum_numero_alterado(db):
    """Linha com 248 no ICV oficial e 132 na grade importada (66 TP + 66
    TS, o mesmo achado real do prompt) devolve a divergência no payload —
    e nenhum dos dois números é alterado pelo cálculo."""
    hoje = date(2026, 8, 19)
    tipo_dia = _tipo_dia(hoje)
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="LINHA_DIVERGENTE", data_referencia=hoje,
        programadas=248, realizadas_tp_ts=103, realizadas_ts_tp=97,
    ))
    db.commit()
    for i in range(66):
        db.add(PartidaProgramada(
            linha_codigo="LINHA_DIVERGENTE", tipo_dia=tipo_dia, numero_tabela=1,
            sequencia=i + 1, terminal="TP", horario=time(4, 10), vigencia=hoje,
        ))
        db.add(PartidaProgramada(
            linha_codigo="LINHA_DIVERGENTE", tipo_dia=tipo_dia, numero_tabela=1,
            sequencia=i + 1, terminal="TS", horario=time(4, 10), vigencia=hoje,
        ))
    db.commit()

    resultado = ranking_prioridade(db, hoje)
    item = next(r for r in resultado if r["linha_codigo"] == "LINHA_DIVERGENTE")

    assert item["programadas_oficial"] == 248  # ⛔ não foi alterado
    assert item["divergencia_denominador"] == 248 - 132

    with Session(db.get_bind()) as verificacao:
        grade = verificacao.execute(
            select(PartidaProgramada).where(PartidaProgramada.linha_codigo == "LINHA_DIVERGENTE")
        ).scalars().all()
    assert len(grade) == 132  # ⛔ a grade também não foi alterada


def test_d20_duas_fontes_lado_a_lado_sem_coluna_combinada(db):
    """Linha com oficial (planilha) e campo (fiscal) no mesmo dia devolve
    os dois separados; nenhuma chave combinada tipo 'icv' existe."""
    hoje = date(2026, 8, 19)
    tipo_dia = _tipo_dia(hoje)
    func = Funcionario(id=uuid4(), re="70030", nome="Fiscal Campo Teste", status="ATIVO")
    ponto = Ponto(codigo="P1", nome="Ponto Teste", terminal="TP", ativo=True)
    db.add_all([func, ponto])
    db.commit()

    db.add(IcvApurado(
        id=uuid4(), linha_codigo="X-10", data_referencia=hoje,
        programadas=100, realizadas_tp_ts=45, realizadas_ts_tp=40,
    ))
    for i in range(50):
        db.add(PartidaProgramada(
            linha_codigo="X-10", tipo_dia=tipo_dia, numero_tabela=1,
            sequencia=i + 1, terminal="TP", horario=time(8, 0), vigencia=hoje,
        ))
    turno = Turno(
        id=uuid4(), funcionario_id=func.id, fiscal_re=func.re, ponto_codigo="P1",
        terminal="TP", periodo="1", data_referencia=hoje, tipo_dia=tipo_dia, status="ABERTO",
    )
    db.add(turno)
    db.commit()
    for i in range(30):
        db.add(RegistroPartida(
            id=uuid4(), turno_id=turno.id, linha_codigo="X-10", numero_tabela=1, terminal="TP",
            horario_programado=time(8, i % 60), resultado="REALIZADA",
        ))
    db.commit()

    resultado = calcular_icv_linha_dia(db, "X-10", hoje)
    assert resultado["programadas_oficial"] == 100
    assert resultado["realizadas_oficial"] == 85
    assert resultado["icv_oficial"] == 85.0
    assert resultado["programadas_campo"] == 50
    assert resultado["realizadas_campo"] == 30
    assert resultado["icv_campo"] == 60.0
    assert "icv" not in resultado  # ⛔ nenhuma coluna combinada


def test_d30_icv_acima_de_100_por_cento_sem_teto(db):
    """programadas=20, realizadas=22 → 110%, sem teto em nenhum lugar."""
    hoje = date(2026, 8, 19)
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="Y-10", data_referencia=hoje,
        programadas=20, realizadas_tp_ts=12, realizadas_ts_tp=10,
    ))
    db.commit()

    resultado = calcular_icv_linha_dia(db, "Y-10", hoje)
    assert resultado["realizadas_oficial"] == 22
    assert resultado["icv_oficial"] == 110.0


# ============================================================================
# BLOCO 4 — painel do coordenador: cascata (D24), RBAC dos novos endpoints
# de ICV e ações da coordenação (D26)
# ============================================================================

def _abrir_turno_e_perder(db, *, func_id, re, ponto_codigo, linha_codigo, data_referencia, tipo_dia, horarios):
    """Helper só destes testes — cria ponto/turno e marca cada horário da
    lista como PERDIDA (motivo OUTRO, sem exigir prefixo/RE)."""
    if db.get(Ponto, ponto_codigo) is None:
        db.add(Ponto(codigo=ponto_codigo, nome="Ponto Teste", terminal="TP", ativo=True))
        db.commit()
    turno = Turno(
        id=uuid4(), funcionario_id=func_id, fiscal_re=re, ponto_codigo=ponto_codigo,
        terminal="TP", periodo="1", data_referencia=data_referencia, tipo_dia=tipo_dia, status="ABERTO",
    )
    db.add(turno)
    db.commit()
    for idx, horario in enumerate(horarios):
        db.add(RegistroPartida(
            id=uuid4(), turno_id=turno.id, linha_codigo=linha_codigo, numero_tabela=idx + 1,
            terminal="TP", horario_programado=horario, resultado="PERDIDA",
            motivo="OUTRO", motivo_outro="teste",
        ))
    db.commit()
    return turno


def test_d24_cascata_duas_perdas_mesma_faixa_acende(db):
    hoje = date(2026, 8, 19)
    func = Funcionario(id=uuid4(), re="70040", nome="Fiscal Cascata", status="ATIVO")
    db.add(func)
    db.commit()
    _abrir_turno_e_perder(
        db, func_id=func.id, re=func.re, ponto_codigo="P1", linha_codigo="1726-10",
        data_referencia=hoje, tipo_dia="UTIL", horarios=[time(18, 10), time(18, 40)],
    )

    from app.services.icv import detectar_cascata
    resultado = detectar_cascata(db, hoje)
    assert len(resultado) == 1
    assert resultado[0] == {"linha_codigo": "1726-10", "faixa_hora": 18, "quantidade": 2}


def test_d24_cascata_faixas_diferentes_nao_acende(db):
    """Uma perda na faixa 18 e outra na 19 NÃO acendem — precisa de 2+ na
    MESMA faixa horária."""
    hoje = date(2026, 8, 19)
    func = Funcionario(id=uuid4(), re="70041", nome="Fiscal Cascata 2", status="ATIVO")
    db.add(func)
    db.commit()
    _abrir_turno_e_perder(
        db, func_id=func.id, re=func.re, ponto_codigo="P1", linha_codigo="1726-10",
        data_referencia=hoje, tipo_dia="UTIL", horarios=[time(18, 50), time(19, 5)],
    )

    from app.services.icv import detectar_cascata
    resultado = detectar_cascata(db, hoje)
    assert resultado == []


def test_icv_ranking_fiscal_nega_403(ambiente):
    _como_fiscal(ambiente)
    resp = ambiente["http"].get("/fiscalizacao/icv/ranking")
    assert resp.status_code == 403, resp.text


def test_icv_cascata_coordenador_acessa(ambiente):
    _como_coordenador(ambiente)
    resp = ambiente["http"].get("/fiscalizacao/icv/cascata", params={"data": "2026-08-19"})
    assert resp.status_code == 200, resp.text
    assert resp.json() == []


def test_acao_coordenacao_fiscal_nega_403(ambiente):
    _como_fiscal(ambiente)
    resp = ambiente["http"].post("/fiscalizacao/acoes", json={
        "linha_codigo": "1726-10", "data_referencia": "2026-08-19",
        "faixa_hora": 18, "descricao": "Reforço na equipe.",
    })
    assert resp.status_code == 403, resp.text


def test_acao_coordenacao_criar_e_listar(ambiente):
    _como_coordenador(ambiente)
    criar = ambiente["http"].post("/fiscalizacao/acoes", json={
        "linha_codigo": "1726-10", "data_referencia": "2026-08-19",
        "faixa_hora": 18, "descricao": "Equipe reforçou o TP às 18h.",
        "resultado_observado": "ICV subiu de 94,62% para 99,46% em dois dias.",
    })
    assert criar.status_code == 201, criar.text
    assert criar.json()["registrado_por"] == str(_COORDENADOR.id)

    listar = ambiente["http"].get("/fiscalizacao/acoes", params={"linha_codigo": "1726-10"})
    assert listar.status_code == 200, listar.text
    assert len(listar.json()) == 1
    assert listar.json()[0]["descricao"] == "Equipe reforçou o TP às 18h."


# ============================================================================
# BLOCO 5 — placar impresso por linha (§7): código, ICV da semana
# anterior, meta ao lado, evolução de 7 dias. ⛔ Nenhum dado de pessoa.
# ============================================================================

def test_placar_evolucao_sete_dias_e_semana_anterior(db):
    hoje = date(2026, 8, 21)
    db.add(Bacia(codigo="B1", nome="Bacia Teste", meta_icv=98.0))
    db.commit()
    db.add(BaciaLinha(bacia_codigo="B1", linha_codigo="1726-10", vigencia_inicio=date(2026, 1, 1)))

    # Hoje e 7 dias atrás (a "semana anterior") têm dado; o resto do
    # intervalo de 7 dias fica sem dado (icv None), pra provar que a
    # ausência não quebra a montagem.
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="1726-10", data_referencia=hoje,
        programadas=100, realizadas_tp_ts=45, realizadas_ts_tp=40,  # 85%
    ))
    db.add(IcvApurado(
        id=uuid4(), linha_codigo="1726-10", data_referencia=hoje - timedelta(days=7),
        programadas=100, realizadas_tp_ts=47, realizadas_ts_tp=45,  # 92%
    ))
    db.commit()

    placar = montar_placar_linha(db, "1726-10", hoje)

    assert placar["linha_codigo"] == "1726-10"
    assert placar["meta_icv"] == 98.0
    assert placar["icv_semana_anterior"] == 92.0
    assert len(placar["evolucao"]) == 7
    assert placar["evolucao"][-1]["data_referencia"] == hoje
    assert placar["evolucao"][-1]["icv"] == 85.0
    assert placar["evolucao"][-1]["fonte"] == "OFICIAL"
    # Dias sem dado nenhum (nem oficial nem grade importada) saem com
    # icv=None, não zero — zero mentiria que o dia foi ruim.
    dias_sem_dado = [d for d in placar["evolucao"][:-1] if d["icv"] is None]
    assert len(dias_sem_dado) == 6


def test_placar_sem_bacia_meta_nula_sem_erro(db):
    """Linha sem bacia cadastrada — meta_icv sai None, não quebra."""
    hoje = date(2026, 8, 21)
    placar = montar_placar_linha(db, "SEM-BACIA", hoje)
    assert placar["meta_icv"] is None
    assert placar["icv_semana_anterior"] is None
    assert len(placar["evolucao"]) == 7


def test_placar_nenhum_dado_pessoal_nas_chaves(db):
    """⛔ Sem nome, RE ou qualquer identificação de pessoa (§7) — nem
    como chave do payload."""
    hoje = date(2026, 8, 21)
    placar = montar_placar_linha(db, "1726-10", hoje)
    chaves = set(placar.keys()) | {k for dia in placar["evolucao"] for k in dia.keys()}
    proibidas = {"re", "nome", "fiscal", "fiscal_re", "coordenador", "motorista_re", "cobrador_re"}
    assert chaves.isdisjoint(proibidas)
