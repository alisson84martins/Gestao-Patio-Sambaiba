"""Lógica de parser e importação de planilha Excel de escala.

Suporta dois formatos automaticamente:

  1. Formato Sambaíba (V2) — múltiplas abas com grupos de colunas:
       Aba E2:      3 grupos (carro|hora|linha) em cols 0,1,2 / 6,7,8 / 12,13,14
                    Cabeçalho em 3 linhas, dados a partir da linha 4
       Aba AR2:     mesma estrutura da E2
       Aba MANOBRA: 6 grupos (carro|hora) em cols 0,1 / 3,4 / 6,7 / 9,10 / 12,13 / 15,16
                    Cabeçalho em 1 linha, dados a partir da linha 2

  2. Formato simples — uma aba, uma linha por ônibus:
       Col A: numero_frota | B: linha_codigo | C: horario_saida | D: re_motorista | E: tipo
       Cabeçalho na linha 1, dados a partir da linha 2

Auto-detecta o formato pelo nome/cabeçalho das abas.
Auto-cria Ônibus e Linha caso não existam no banco.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from datetime import date as date_type, datetime, time, timezone
from typing import IO

from openpyxl import load_workbook
from sqlalchemy import select, update
from sqlalchemy.orm import Session

from app.models import (
    Escala,
    ImportacaoEscala,
    Linha,
    Motorista,
    Onibus,
    OrigemEscalaEnum,
    StatusImportacaoEnum,
    TipoEscalaEnum,
)
from app.models.enums import SetorEnum


# ─── Dataclass de linha parseada ─────────────────────────────────────────────

@dataclass
class LinhaParseada:
    linha_planilha: int
    numero_frota: int | None = None
    linha_codigo: str | None = None
    horario_saida: time | None = None
    re_motorista: str | None = None
    tipo: TipoEscalaEnum | None = None
    erro: str | None = None


# ─── Helpers de parsing ───────────────────────────────────────────────────────

def _parse_horario(valor) -> time | None:
    if valor is None:
        return None
    if isinstance(valor, time):
        return valor
    if isinstance(valor, datetime):
        return valor.time()
    if isinstance(valor, str):
        s = valor.strip()
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                return datetime.strptime(s, fmt).time()
            except ValueError:
                continue
    return None


def _parse_tipo(valor) -> TipoEscalaEnum | None:
    if not valor:
        return None
    try:
        return TipoEscalaEnum(str(valor).strip().upper())
    except ValueError:
        return None


def _val_carro(valor) -> int | None:
    """Valida número de frota (3-4 dígitos). Retorna None se inválido."""
    if valor is None:
        return None
    try:
        s = str(int(round(float(str(valor))))).strip()
    except (ValueError, TypeError):
        return None
    return int(s) if re.match(r'^\d{3,4}$', s) else None


def _detectar_tipo_aba(sheet_name: str, header_text: str) -> str | None:
    """Detecta tipo da aba: 'e2', 'ar2', 'manobra', 'configuracao' ou None."""
    n = sheet_name.upper()
    h = header_text.upper()
    if 'E2' in n or 'E2' in h:
        return 'e2'
    if 'AR2' in n or 'AR2' in h:
        return 'ar2'
    if any(k in n or k in h for k in ('MANOBRA', 'MANOBRISTA', 'PRESO')):
        return 'manobra'
    if 'CONFIGURA' in n or 'CONFIGURA' in h:
        return 'configuracao'
    return None


def _setor_por_frota(numero_frota: int) -> SetorEnum:
    """Infere setor pelo número de frota: 1000-1999 = E2; 2000-2999 = AR2."""
    if 2000 <= numero_frota <= 2999:
        return SetorEnum.AR2
    return SetorEnum.E2  # default e para 1xxx


# ─── Parser formato Sambaíba (V2) ────────────────────────────────────────────

def _parsear_formato_sambaiba(wb) -> list[LinhaParseada]:
    """
    Parseia o formato real das planilhas de escala da Sambaíba.

    Abas E2/AR2:  3 grupos (carro, hora, linha) por linha de dados
                  Índices: (0,1,2), (6,7,8), (12,13,14)
                  Dados a partir da linha 4 (3 linhas de cabeçalho)

    Aba MANOBRA:  6 grupos (carro, hora) por linha de dados
                  Índices: (0,1), (3,4), (6,7), (9,10), (12,13), (15,16)
                  Dados a partir da linha 2 (1 linha de cabeçalho)
    """
    resultado: list[LinhaParseada] = []
    idx = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Header = primeiras 3 linhas concatenadas
        header_text = ' '.join(str(c or '') for row in rows[:3] for c in row)
        tipo_aba = _detectar_tipo_aba(sheet_name, header_text)

        if not tipo_aba or tipo_aba == 'configuracao':
            continue

        if tipo_aba == 'e2':
            tipo_enum = TipoEscalaEnum.PLANTAO_E2
            grupos = [(0, 1, 2), (6, 7, 8), (12, 13, 14)]
            data_rows = rows[3:]  # pula 3 linhas de cabeçalho
        elif tipo_aba == 'ar2':
            tipo_enum = TipoEscalaEnum.PLANTAO_AR2
            grupos = [(0, 1, 2), (6, 7, 8), (12, 13, 14)]
            data_rows = rows[3:]
        else:  # manobra
            tipo_enum = TipoEscalaEnum.MANOBRA
            grupos = [(0, 1), (3, 4), (6, 7), (9, 10), (12, 13), (15, 16)]
            data_rows = rows[1:]  # pula 1 linha de cabeçalho

        for row in data_rows:
            if not row or all(c is None for c in row):
                continue
            # Garante largura mínima para evitar IndexError
            row_list = list(row) + [None] * 20

            for grupo in grupos:
                ci = grupo[0]
                hi = grupo[1]
                li = grupo[2] if len(grupo) > 2 else None

                frota = _val_carro(row_list[ci])
                if frota is None:
                    continue

                hora_raw = row_list[hi]
                hora_str = str(hora_raw or '').strip().upper()
                # Marcadores especiais não são escalas de serviço
                if hora_str in ('PRESO', 'AMOSTRAL', 'EVENTO', 'TABELAS', ''):
                    continue

                hora = _parse_horario(hora_raw)
                if hora is None:
                    continue  # hora não reconhecida — pula silenciosamente

                linha_codigo: str | None = None
                if li is not None and row_list[li] is not None:
                    linha_codigo = str(row_list[li]).strip() or None

                idx += 1
                l = LinhaParseada(
                    linha_planilha=idx,
                    numero_frota=frota,
                    linha_codigo=linha_codigo,
                    horario_saida=hora,
                    tipo=tipo_enum,
                )

                # Plantão SEM linha = erro (manobra não exige linha)
                if tipo_enum != TipoEscalaEnum.MANOBRA and not linha_codigo:
                    l.erro = "linha_codigo ausente na aba de plantão"

                resultado.append(l)

    return resultado


# ─── Parser formato simples ───────────────────────────────────────────────────

def _parsear_formato_simples(wb) -> list[LinhaParseada]:
    """Formato simples: uma aba, col A=frota, B=linha, C=hora, D=re, E=tipo."""
    ws = wb.active
    linhas: list[LinhaParseada] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(cell is None for cell in row):
            continue
        l = LinhaParseada(linha_planilha=idx)
        try:
            cells = list(row) + [None] * 5
            l.numero_frota = int(cells[0]) if cells[0] is not None else None
            l.linha_codigo = str(cells[1]).strip() if cells[1] is not None else None
            l.horario_saida = _parse_horario(cells[2])
            l.re_motorista = str(cells[3]).strip() if cells[3] is not None else None
            l.tipo = _parse_tipo(cells[4])
            if l.numero_frota is None:
                l.erro = "numero_frota vazio"
            elif l.horario_saida is None:
                l.erro = "horario_saida inválido"
        except Exception as exc:  # noqa: BLE001
            l.erro = f"erro ao parsear: {exc}"
        linhas.append(l)
    return linhas


# ─── Auto-detecção de formato ─────────────────────────────────────────────────

def _e_formato_sambaiba(wb) -> bool:
    """True se qualquer aba tiver nome/cabeçalho reconhecido como E2/AR2/Manobra."""
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        header_text = ' '.join(str(c or '') for row in rows for c in row)
        tipo = _detectar_tipo_aba(sheet_name, header_text)
        if tipo and tipo != 'configuracao':
            return True
    return False


def parsear_planilha(stream: IO[bytes]) -> list[LinhaParseada]:
    """Ponto de entrada único: auto-detecta formato e delega ao parser correto."""
    wb = load_workbook(stream, data_only=True, read_only=True)
    if _e_formato_sambaiba(wb):
        return _parsear_formato_sambaiba(wb)
    return _parsear_formato_simples(wb)


# ─── Auto-criação de catálogos ────────────────────────────────────────────────

def _get_or_create_onibus(db: Session, numero_frota: int) -> Onibus:
    """Busca ou cria ônibus pelo número de frota."""
    onibus = db.execute(
        select(Onibus).where(Onibus.numero_frota == numero_frota)
    ).scalar_one_or_none()
    if onibus is None:
        onibus = Onibus(numero_frota=numero_frota)
        db.add(onibus)
        db.flush()
    return onibus


def _get_or_create_linha(db: Session, codigo: str, setor: SetorEnum) -> Linha:
    """Busca ou cria linha pelo código. Se criar, usa código como nome e setor inferido."""
    linha = db.execute(
        select(Linha).where(Linha.codigo == codigo)
    ).scalar_one_or_none()
    if linha is None:
        linha = Linha(
            codigo=codigo,
            nome=codigo,   # nome provisório = código; pode ser atualizado depois
            setor=setor,
            ativa=True,
        )
        db.add(linha)
        db.flush()
    return linha


def hash_arquivo(conteudo: bytes) -> str:
    return hashlib.sha256(conteudo).hexdigest()


# ─── Importação principal ─────────────────────────────────────────────────────

def importar_escala(
    db: Session,
    arquivo_nome: str,
    conteudo: bytes,
    data_escala: date_type,
    tipo_default: TipoEscalaEnum,
    importado_por_id,
    substituir_existentes: bool = True,
) -> tuple[ImportacaoEscala, list[dict], int]:
    """Processa o Excel e insere escalas. Retorna (importacao, erros, substituidas)."""
    import io
    linhas_lidas = parsear_planilha(io.BytesIO(conteudo))

    # Soft-delete das escalas anteriores da mesma data
    substituidas = 0
    if substituir_existentes:
        stmt = (
            update(Escala)
            .where(Escala.data == data_escala, Escala.deletado_em.is_(None))
            .values(deletado_em=datetime.now(timezone.utc))
        )
        result = db.execute(stmt)
        substituidas = result.rowcount or 0

    # Registro de importação
    imp = ImportacaoEscala(
        arquivo_nome=arquivo_nome,
        arquivo_hash=hash_arquivo(conteudo),
        data_escala=data_escala,
        total_registros=len(linhas_lidas),
        status=StatusImportacaoEnum.PARCIAL,
        importado_por=importado_por_id,
    )
    db.add(imp)
    db.flush()

    # Caches para evitar queries repetidas por ônibus/linha/motorista
    onibus_cache: dict[int, Onibus] = {}
    linha_cache: dict[str, Linha] = {}
    motorista_cache: dict[str, Motorista] = {}

    erros: list[dict] = []
    sucesso = 0

    for l in linhas_lidas:
        if l.erro:
            erros.append({
                "linha": l.linha_planilha,
                "motivo": l.erro,
                "valor_recebido": None,
            })
            continue

        tipo_efetivo = l.tipo or tipo_default

        # ── Ônibus (auto-cria) ────────────────────────────────────────────────
        onibus = onibus_cache.get(l.numero_frota)
        if onibus is None:
            try:
                onibus = _get_or_create_onibus(db, l.numero_frota)
                onibus_cache[l.numero_frota] = onibus
            except Exception as exc:
                erros.append({
                    "linha": l.linha_planilha,
                    "motivo": f"Erro ao registrar ônibus {l.numero_frota}: {exc}",
                    "valor_recebido": str(l.numero_frota),
                })
                continue

        # ── Linha (auto-cria; MANOBRA sem linha usa placeholder "MAN-<setor>") ─
        linha_codigo = l.linha_codigo
        if not linha_codigo:
            # Manobra sem código: usa placeholder por setor
            setor_frota = _setor_por_frota(l.numero_frota)
            linha_codigo = f"MAN-{setor_frota.value}"

        linha = linha_cache.get(linha_codigo)
        if linha is None:
            try:
                # Setor do tipo de escala tem prioridade; fallback pelo frota
                if tipo_efetivo == TipoEscalaEnum.PLANTAO_E2:
                    setor = SetorEnum.E2
                elif tipo_efetivo == TipoEscalaEnum.PLANTAO_AR2:
                    setor = SetorEnum.AR2
                else:
                    setor = _setor_por_frota(l.numero_frota)
                linha = _get_or_create_linha(db, linha_codigo, setor)
                linha_cache[linha_codigo] = linha
            except Exception as exc:
                erros.append({
                    "linha": l.linha_planilha,
                    "motivo": f"Erro ao registrar linha {linha_codigo}: {exc}",
                    "valor_recebido": linha_codigo,
                })
                continue

        # ── Motorista (opcional) ──────────────────────────────────────────────
        motorista_id = None
        if l.re_motorista:
            motorista = motorista_cache.get(l.re_motorista)
            if motorista is None:
                motorista = db.execute(
                    select(Motorista).where(Motorista.re == l.re_motorista)
                ).scalar_one_or_none()
                if motorista:
                    motorista_cache[l.re_motorista] = motorista
            if motorista:
                motorista_id = motorista.id

        # ── Insere escala ─────────────────────────────────────────────────────
        escala = Escala(
            data=data_escala,
            onibus_id=onibus.id,
            motorista_id=motorista_id,
            linha_id=linha.id,
            horario_saida=l.horario_saida,
            tipo=tipo_efetivo,
            origem=OrigemEscalaEnum.IMPORTACAO_EXCEL,
            importacao_id=imp.id,
            criado_por=importado_por_id,
        )
        db.add(escala)
        sucesso += 1

    # ── Atualiza status da importação ─────────────────────────────────────────
    imp.registros_sucesso = sucesso
    imp.registros_erro = len(erros)
    if erros and sucesso == 0:
        imp.status = StatusImportacaoEnum.ERRO
    elif erros:
        imp.status = StatusImportacaoEnum.PARCIAL
    else:
        imp.status = StatusImportacaoEnum.SUCESSO
    if erros:
        imp.erro_detalhe = "\n".join(
            f"Linha {e['linha']}: {e['motivo']}" for e in erros[:10]
        )

    db.commit()
    db.refresh(imp)
    return imp, erros, substituidas
