"""Importação da grade a partir da escala gerencial (.xlsx) — Bloco C.

Porta síncrona do parser de julho (_handoff-claude/Coordenadoria_Sambaiba/
analise-pre-juncao/Escalas Gerenciais/backend-fiscal/app/routers/escalas.py
:: _parse_aba, _serial_to_time, _parse_range_second, _is_ts_label) — ~400
linhas de heurística sobre uma planilha muito irregular, já provadas contra
o arquivo real (Pasta1.xlsx, aba 1726-10: 9 tabelas, 66 partidas TP + 66
TS). Reescrever do zero jogaria fora trabalho que funciona (§7.2 do prompt).

🆕 Estendido (§7.3) para carimbar o período (D8) a partir das linhas
"TERM. 1º PERIODO" / "INICIO 2º PERIODO", que vêm DEPOIS da seção DUPLAS —
o parser de julho parava em DUPLAS e nunca lia o que vem depois. ⛔ Não
chuta: se os rótulos não forem encontrados na aba, ou a hora cair entre os
dois limites, grava periodo = NULL e o endpoint conta quantas ficaram assim.

A seção DUPLAS não traz nome nem chapa de ninguém (confirmado no arquivo
real, §7.4) — este módulo não implementa pré-preenchimento de dupla.
"""
import io
import re
from datetime import date, time
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.models.fiscalizacao import PartidaProgramada

_RE_ABA = re.compile(r'^[0-9A-Z]{4,5}-\d{2}$')

_ROTULO_TERM_1 = "TERM. 1º PERIODO"
_ROTULO_INICIO_2 = "INICIO 2º PERIODO"


def _serial_to_time(val) -> Optional[time]:
    """Converte valor Excel em time. Aceita datetime.time ou fração de dia (0–1)."""
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, float):
        if val >= 1.0:
            return None  # data/número (km, horas trabalhadas), não horário
        total_minutes = round(val * 24 * 60)
        h, m = divmod(total_minutes, 60)
        return time(h % 24, m)
    return None


def _parse_range_second(val: str) -> Optional[time]:
    """Extrai o SEGUNDO horário de uma string 'HH:MM-HH:MM'. Usado para
    células de refeição: '07:55-08:55' → 08:55 (partida após refeição)."""
    try:
        partes = val.strip().split("-")
        if len(partes) == 2:
            h_m = partes[1].strip().split(":")
            if len(h_m) == 2:
                return time(int(h_m[0]) % 24, int(h_m[1]))
    except (ValueError, IndexError):
        pass
    return None


def _is_ts_label(col_a) -> bool:
    """True se o rótulo da coluna A indica terminal TS (Metrô / terminal secundário)."""
    if not isinstance(col_a, str):
        return False
    label = col_a.strip().upper()
    return "METR" in label or label == "TS"


def _localizar_cabecalho(rows: list[tuple]) -> Optional[int]:
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip().upper() == "TABELAS":
            return i
    return None


def _mapear_colunas_tabela(header_row: tuple) -> dict[int, int]:
    tabela_cols: dict[int, int] = {}
    for j, cell in enumerate(header_row):
        if j == 0:
            continue
        if isinstance(cell, (int, float)) and 1 <= int(cell) <= 30:
            tabela_cols[j] = int(cell)
    return tabela_cols


def _coletar_data_rows(rows: list[tuple], header_row_idx: int) -> list[tuple]:
    """Pula HORÁRIOS e linhas em branco; PARA em DUPLAS (seção de escala
    de pessoal — não é horário)."""
    data_rows: list = []
    for row in rows[header_row_idx + 1:]:
        if not row:
            continue
        col_a = row[0]
        if isinstance(col_a, str) and "DUPLAS" in col_a.strip().upper():
            break
        if col_a is None:
            continue
        if isinstance(col_a, str) and col_a.strip().upper() == "HORÁRIOS":
            continue
        data_rows.append(row)
    return data_rows


def _parse_aba(ws: Worksheet) -> list[dict]:
    """Parseia uma aba da escala Sambaíba — porta 1:1 do parser de julho.

    Estrutura real das linhas de dados:
      HORÁRIOS      | 04:15 | ...  ← saída da garagem, IGNORAR
      CEM PQ (TP)   | 04:30 | ...  ← partida do TP   ┐ par
      METRÔ (TS)    | 05:20 | ...  ← partida do TS   ┘
      CEM PQ (TP)   | '07:55-08:55' | ...  ← chegada 07:55, saída 08:55 (refeição) → 2º horário
      CEM PQ (TP)   | 13:10 | ...  ← CHEGADA ao TP (METRÔ seguinte tem 'TERM.') → IGNORAR
      DUPLAS        | ...          ← seção de informações — PARAR aqui

    Regra do par (CEM PQ, METRÔ): se o METRÔ da coluna tem horário real →
    CEM PQ é PARTIDA TP + METRÔ é PARTIDA TS. Se o METRÔ tem string
    ('TERM.', 'REC.', 'INICIO') ou None → CEM PQ era CHEGADA → ignora o par.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_idx = _localizar_cabecalho(rows)
    if header_row_idx is None:
        return []

    tabela_cols = _mapear_colunas_tabela(rows[header_row_idx])
    if not tabela_cols:
        return []

    data_rows = _coletar_data_rows(rows, header_row_idx)

    partidas: list[dict] = []
    seq_tp: dict[int, int] = {t: 0 for t in tabela_cols.values()}
    seq_ts: dict[int, int] = {t: 0 for t in tabela_cols.values()}

    i = 0
    while i < len(data_rows):
        row = data_rows[i]
        col_a = row[0]

        if not isinstance(col_a, str):
            i += 1
            continue

        if not _is_ts_label(col_a):
            # Linha TP — consome este e o próximo (TS)
            ts_row = data_rows[i + 1] if i + 1 < len(data_rows) else None

            for col_idx, num_tabela in tabela_cols.items():
                tp_cell = row[col_idx] if col_idx < len(row) else None
                ts_cell = ts_row[col_idx] if (ts_row and col_idx < len(ts_row)) else None

                ts_time = _serial_to_time(ts_cell)
                if ts_time is None:
                    # TS tem string marcador ('TERM.', 'REC.', 'INICIO') ou
                    # None → TP era chegada, não partida → ignora o par.
                    continue

                tp_time = _serial_to_time(tp_cell)
                if tp_time is None and isinstance(tp_cell, str) and "-" in tp_cell:
                    tp_time = _parse_range_second(tp_cell)

                if tp_time is not None:
                    seq_tp[num_tabela] += 1
                    partidas.append({
                        "numero_tabela": num_tabela, "sequencia": seq_tp[num_tabela],
                        "horario": tp_time, "terminal": "TP",
                    })

                seq_ts[num_tabela] += 1
                partidas.append({
                    "numero_tabela": num_tabela, "sequencia": seq_ts[num_tabela],
                    "horario": ts_time, "terminal": "TS",
                })

            i += 2  # consome TP + TS
        else:
            # Linha TS sem TP precedente (não deve ocorrer normalmente)
            i += 1

    return partidas


def _parse_limites_periodo(
    rows: list[tuple], tabela_cols: dict[int, int]
) -> dict[int, tuple[Optional[time], Optional[time]]]:
    """§7.3 — carimba o período a partir de TERM. 1º PERIODO / INICIO 2º
    PERIODO, linhas que vêm DEPOIS da seção DUPLAS (o parser de julho
    parava antes e nunca lia isso). Devolve, por tabela, (term_1, inicio_2).
    """
    limites: dict[int, tuple[Optional[time], Optional[time]]] = {
        t: (None, None) for t in tabela_cols.values()
    }
    for row in rows:
        col_a = row[0]
        if not isinstance(col_a, str):
            continue
        rotulo = col_a.strip().upper()
        if rotulo == _ROTULO_TERM_1:
            for col_idx, num_tabela in tabela_cols.items():
                valor = row[col_idx] if col_idx < len(row) else None
                _, inicio2 = limites[num_tabela]
                limites[num_tabela] = (_serial_to_time(valor), inicio2)
        elif rotulo == _ROTULO_INICIO_2:
            for col_idx, num_tabela in tabela_cols.items():
                valor = row[col_idx] if col_idx < len(row) else None
                term1, _ = limites[num_tabela]
                limites[num_tabela] = (term1, _serial_to_time(valor))
    return limites


def _classificar_periodo(horario: time, limites: tuple[Optional[time], Optional[time]]) -> Optional[str]:
    """horario <= TERM. 1º PERIODO → '1'; horario >= INICIO 2º PERIODO →
    '2'; senão (ou rótulo não encontrado na aba) → NULL. ⛔ Não chuta."""
    term1, inicio2 = limites
    if term1 is not None and horario <= term1:
        return "1"
    if inicio2 is not None and horario >= inicio2:
        return "2"
    return None


def importar_escala(db: Session, conteudo: bytes, *, tipo_dia: str, vigencia: date) -> dict:
    """Recebe os bytes do .xlsx da escala gerencial. Para cada aba com nome
    no padrão XXXX-YY: apaga as partidas antigas do mesmo tipo_dia+vigência
    para aquela linha e insere as novas, já com o período carimbado (§7.3).

    Responde com um resumo (§7.1): linhas processadas, partidas criadas,
    quantas ficaram com periodo NULL, e os erros — uma aba ruim não derruba
    o upload inteiro, só entra na lista de erros.
    """
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)

    resumo_linhas: list[dict] = []
    erros: list[str] = []
    total_partidas = 0
    total_periodo_nulo = 0

    for sheet_name in wb.sheetnames:
        codigo = sheet_name.strip().upper().replace(" ", "")
        if not _RE_ABA.match(codigo):
            continue

        ws = wb[sheet_name]
        try:
            partidas = _parse_aba(ws)
        except Exception as exc:  # planilha real é muito irregular — uma
            # aba ruim não pode abortar as outras.
            erros.append(f"{codigo}: falha ao interpretar a aba ({exc})")
            continue

        if not partidas:
            continue

        rows = list(ws.iter_rows(values_only=True))
        header_row_idx = _localizar_cabecalho(rows)
        tabela_cols = _mapear_colunas_tabela(rows[header_row_idx]) if header_row_idx is not None else {}
        limites_por_tabela = _parse_limites_periodo(rows, tabela_cols) if tabela_cols else {}

        db.execute(
            delete(PartidaProgramada).where(
                PartidaProgramada.linha_codigo == codigo,
                PartidaProgramada.tipo_dia == tipo_dia,
                PartidaProgramada.vigencia == vigencia,
            )
        )
        db.flush()  # garante que o DELETE foi enviado antes do INSERT (armadilha já registrada no projeto)

        periodo_nulo_aba = 0
        for p in partidas:
            limites = limites_por_tabela.get(p["numero_tabela"], (None, None))
            periodo = _classificar_periodo(p["horario"], limites)
            if periodo is None:
                periodo_nulo_aba += 1
            db.add(PartidaProgramada(
                linha_codigo=codigo,
                tipo_dia=tipo_dia,
                numero_tabela=p["numero_tabela"],
                sequencia=p["sequencia"],
                terminal=p["terminal"],
                horario=p["horario"],
                periodo=periodo,
                vigencia=vigencia,
            ))

        total_partidas += len(partidas)
        total_periodo_nulo += periodo_nulo_aba
        resumo_linhas.append({
            "linha": codigo,
            "partidas_importadas": len(partidas),
            "periodo_nulo": periodo_nulo_aba,
        })

    return {
        "vigencia": vigencia.isoformat(),
        "tipo_dia": tipo_dia,
        "linhas": resumo_linhas,
        "total_partidas": total_partidas,
        "total_periodo_nulo": total_periodo_nulo,
        "erros": erros,
    }
